import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
import datetime as dt
import logging

from ALGO.excel_formatting_module import ExcelFormatting

logger = logging.getLogger(__name__)
# added logging statements

"""
Replace the government bond yield curve (which is quite wrong to use) and either use the LIBOR
rates or the OIS (Overnight Index Swap) rates
https://www-2.rotman.utoronto.ca/~hull/DownloadablePublications/LIBORvsOIS.pdf
Traditionally practitioners have used LIBOR and LIBOR-swap rates as proxies for risk-free rates
when valuing derivatives. This practice has been called into question by the credit crisis that
started in 2007. Many banks now consider that overnight indexed swap (OIS) rates should be
used as the risk-free rate when collateralized portfolios are valued and that LIBOR should be
used for this purpose when portfolios are not collateralized. This paper examines this practice
and concludes that OIS rates should be used in all situations.
"""


class bondYields:
    def __init__(self, todays_date, cwd):
        logger.debug("Initialization of bond gathering")
        self.bond_yields = {}
        self.todays_date = todays_date
        self.cwd = cwd

    # preferred rate for pricing derivatives
    # the rate at which you can borrow/lend cash is the rate you need to consider
    def overnightindexedswaps(self, days_to_expiry, overnight_rate):
        # the theory behind this is to calculate the spread between a compounded daily index rate (the overnight rate)
        # and a fixed index rate. this is valid up to a two week term

        # first step is to multiply the overnight rate for the period in which the swap applies
        # the overnight lending rate is typically the federal funds rate or the SOFR (secured overnight financing rate)
        floating_rate_leg = ((overnight_rate * days_to_expiry) / 360) + 1

        # fixed rate is the LIBOR rate as it is the least risky rate available
        # this also should pretty much always be larger
        fixed_rate_leg = ((overnight_rate / 360) + 1) ** days_to_expiry

        ois_rate = (fixed_rate_leg - floating_rate_leg) * 100
        logging.debug(f"OIS Rate: {ois_rate}% given overnight rate of {overnight_rate}% "
                      f"with {days_to_expiry} days to expiration")
        return round(ois_rate, 6)

    # secured overnight financing rate is the rate at which banks lend to each other overnight, fixed leg of OIS
    def sofr(self):
        with requests.get("https://fred.stlouisfed.org/series/SOFR", stream=True) as r:
            soup = BeautifulSoup(r.text, 'lxml')
            sofr_list = [entry.text for entry in soup.find_all('span', {'class': 'series-meta-observation-value'})]
            sofr = float(sofr_list[-1])
        logging.debug(f"SOFR rate: {sofr}%")
        return sofr

    # LIBOR rates as used as the floating leg in the OIS calculation
    def LIBOR_yields(self):
        # this is also an option, although i think its much more reliant on website structure to get the specific table
        tables = pd.read_html("https://www.global-rates.com/en/interest-rates/libor/american-dollar/american-dollar.aspx")
        libor_df = None
        for i, table in enumerate(tables):
            # this will sort out the correct dataframe we need
            try:
                libor_df = table.set_index([0], drop=True)
                if libor_df.empty or libor_df.isnull().values.any() or len(libor_df.index) < 2:
                    raise Exception
                libor_df = libor_df.rename(columns=libor_df.iloc[0]).drop(libor_df.index[0])

                dates = list(libor_df.columns)
                # this is a bit cringeworthy but works well, swaps unknown values
                for index, column in libor_df.copy().iteritems():
                    date_int = dates.index(column.name)
                    try:
                        # try and use the right date's value if there is an unknown yield
                        libor_df[dates[date_int]][libor_df[dates[date_int]] == '-'] = libor_df[dates[date_int + 1]]
                    except IndexError:
                        try:
                            libor_df[dates[date_int]][libor_df[dates[date_int]] == '-'] = libor_df[dates[date_int - 1]]
                        except IndexError:
                            libor_df[dates[date_int]][libor_df[dates[date_int]] == '-'] = "0.00000 %"
                        pass

                libor_df[libor_df.columns[0:]] = libor_df[libor_df.columns[0:]].replace('\xa0%', '',
                                                                                        regex=True).astype(float)
                dt.datetime.strptime(libor_df.columns[0], "%m-%d-%Y")
                break
            except Exception as e:
                continue

        date = libor_df.columns[0]

        url = self.cwd + fr"\Daily Stock Analysis\Bonds\LIBOR Yields ({date}).xlsx"
        wb = openpyxl.Workbook()
        wb.save(url)
        book = openpyxl.load_workbook(url)
        writer = pd.ExcelWriter(url, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        libor_df.to_excel(writer, sheet_name=f"LIBOR yields for {date}")
        try:
            sheet = book['Sheet']
            book.remove(sheet)
        except KeyError:
            pass
        writer.save()
        writer.close()
        book.save(url)
        book.close()

        ExcelFormatting(file_path=url).formatting()
        libor_yields = tuple(libor_df[date].to_list())

        logging.debug(f"London Interbank Offering Rates (LIBOR): {libor_yields}")
        return libor_yields

    # note that this the bank discount rate
    def treasury_bond_yields(self):
        month = f"{self.todays_date.month:02d}" if self.todays_date.month < 10 else self.todays_date.month

        bonddf = pd.read_html(f"https://home.treasury.gov/resource-center/data-chart-center/interest-rates/Text"
                              f"View?typdaily_treasury_yield_curve&field_tdr_date_value_"
                              f"month={self.todays_date.year}{month}")[0]

        bonddf['Date'] = bonddf['Date'].str.replace("/", '-')
        bond_df = bonddf.set_index('Date', drop=True)

        bond_df = bond_df[['4 WEEKS BANK DISCOUNT', '13 WEEKS BANK DISCOUNT', '26 WEEKS BANK DISCOUNT',
                          '52 WEEKS BANK DISCOUNT']]

        last_updated_date = bond_df.tail(1).index.item()

        url = self.cwd + f"\\Daily Stock Analysis\\Bonds\\US T-Bond Yields ({last_updated_date}).xlsx"
        wb = openpyxl.Workbook()
        wb.save(url)
        book = openpyxl.load_workbook(url)
        writer = pd.ExcelWriter(url, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        bond_df.to_excel(writer, sheet_name=f"T-Bonds, {last_updated_date}")
        try:
            sheet = book['Sheet']
            book.remove(sheet)
        except KeyError:
            pass
        writer.save()
        writer.close()
        book.save(url)
        book.close()

        ExcelFormatting(file_path=url).formatting()
        logging.debug("Treasury bond excel sheet saved successfully")

        yield_tuple = tuple(bond_df.loc[last_updated_date])

        logging.debug(yield_tuple)

        return yield_tuple
