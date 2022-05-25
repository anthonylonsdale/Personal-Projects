import datetime as dt
import pandas as pd
import ctypes
import yahooquery
import openpyxl
import concurrent.futures
import os
import sqlite3
import logging

from ALGO.excel_formatting_module import ExcelFormatting

logger = logging.getLogger(__name__)


def unique_cols(series):
    a = series.to_numpy()
    return (a[0] == a).all()


class Options:
    def __init__(self, stock_tickers=None, initial_data=None, quote_data=None, rate=None, cwd=None, bbond=None):
        self.stock_tickers = stock_tickers
        self.initial_data = initial_data
        self.quote_data = quote_data
        self.rate = rate
        self.option_value = {}
        self.today = dt.datetime.now()
        self.cwd = cwd
        self.bond_bootstrapper = bbond
        for ticker in stock_tickers:
            self.option_value[ticker] = []

    # work in progress
    def options_fetch(self, stock_quotes):
        max_pain = {}
        strike_pins = {}

        for stock in self.stock_tickers:
            max_pain[stock] = 0
            query = yahooquery.Ticker([stock], asynchronous=True)
            options = query.option_chain

            expiration_dates = list(options.index.unique(level=1))
            connection = sqlite3.connect(f'{self.cwd}\\Databases\\options.db')
            expiration = expiration_dates[0].to_pydatetime().date()

            """
            Underlying theory behind why this is necessary
            Options open interest (and subsequently the total dollar value) is important because options are now the
            predominant driving force behind stock price movements since March of 2020. What I need to do here is
            calculate the value of options that are out of the money (which subsequently will change the stock price
            in that direction) and options that are in the money. We want to assume that the market makers who are 
            selling these options will want to "pin" the stock price to the point where they will profit the most.
            How do I determine this? A large options volume can be a hedge fund trader that is trying to move a 
            stock in the short-term, in which case volume is very important to track throughout the day. However, 
            I will assume that the vast financial resources in institutions is going to eventually push the stock 
            price to the price where the most options contracts will expire.
            """
            call_df = pd.read_sql(f'select * from "{stock} Calls {expiration}"', con=connection).set_index('strike')
            for index, row in call_df.copy().iterrows():
                call_df.at[index, 'max_pain'] = row['openInterest'] * row['lastPrice']

            put_df = pd.read_sql(f'select * from "{stock} Puts {expiration}"', con=connection).set_index('strike')
            for index, row in put_df.copy().iterrows():
                put_df.at[index, 'max_pain'] = row['openInterest'] * row['lastPrice']

            reduced_df = call_df['max_pain'].add(put_df['max_pain'], fill_value=0)

            numerator = 0
            denominator = 0
            for index, row in reduced_df.iteritems():
                numerator += row * index
                denominator += row

            max_pain_strike = numerator / denominator
            max_pain[stock] = round(max_pain_strike, 2)

            # look at options pinning
            # we want to weight each strike with consideration to the max pain
            call_weight_sum = call_df['max_pain'].sum()
            call_df['weight'] = round(call_df['max_pain'] / call_weight_sum, 6)
            put_weight_sum = put_df['max_pain'].sum()
            put_df['weight'] = round(put_df['max_pain'] / put_weight_sum, 6)

            # pick out the strikes that are actually being actively traded (which should mean a greater chance of an
            # effect on price movement i.e. if a particular strike is being traded, then there may be an underlying
            # position that is purchased or short sold based on that options trade)
            # put_df = put_df.loc[put_df['volume'] >= 10]
            put_df = put_df.loc[put_df['weight'] >= .005]
            # call_df = call_df.loc[call_df['volume'] >= 10]
            call_df = call_df.loc[call_df['weight'] >= .005]

            current_price = stock_quotes[stock][-1]['current price']
            # in the money call options can be considered supports as the option holders want their options to stay itm
            support_strikes = call_df.iloc[call_df.index <= current_price]
            support_strikes_dict = pd.Series(support_strikes['weight'], index=support_strikes.index).to_dict()

            # likewise with in the money puts, these can be considered resistances
            resistance_strikes = put_df.iloc[put_df.index >= current_price]
            resistance_strikes_dict = pd.Series(resistance_strikes['weight'], index=resistance_strikes.index).to_dict()

            # out of the money options are tricky to consider so i will ignore them (for now)
            strike_pins[stock] = {'Call Supports': support_strikes_dict, 'Put Resistances': resistance_strikes_dict}

        return max_pain, strike_pins

    # since this module also creates the options database, we need to check in this function and make sure old tables
    # are dropped
    def thread_marshaller(self):
        # ok so at this point we have all old and non-relevant stock tables dropped, we need to check and see if the
        # timestamp table exists since that determines when the last options chain insertion was
        with sqlite3.connect(f'{self.cwd}\\Databases\\options.db') as db:
            db.execute("CREATE TABLE IF NOT EXISTS timestamp (stock text, time timestamp, expiration date, "
                       "table_no integer, unique (table_no))")
            db.commit()

            cur = db.execute("select * from timestamp")
            timestamps = cur.fetchall()
            outdated_options_chains = []

            if not len(timestamps) == 0:
                for element in timestamps:
                    timestamp = dt.datetime.strptime(element[1], "%Y-%m-%d %H:%M:%S.%f")
                    if timestamp < (dt.datetime.now() - dt.timedelta(minutes=15)):
                        outdated_options_chains.append(element[0])
                        logging.debug(f"Options chain for {element[0]} outdated (gathered {timestamp}), replacing.")
                        db.execute("delete from timestamp where table_no = (?)", (self.stock_tickers.index(element[0]),))
                        db.commit()
            else:
                # if there are no timestamps then that means we need to gather all of the options chains
                outdated_options_chains = self.stock_tickers

        logging.debug(f"{outdated_options_chains}: options chains that are either outdated or need to be overwritten")

        if len(outdated_options_chains) == 0:
            logging.debug("No options chains need to be gathered at this point")
            # sometimes the program wipes the excel files containing the options chain because it gets corrupted
            # when I auto format it, usually its because i terminate the program accidentally in the middle of the
            # formatting process or for some other reason. This should be quite rare because i fixed the c# program
            # that formats excel files in order to specifically avoid this, but just in case i have this.
            for stock in self.stock_tickers:
                if not os.path.isfile(f'{self.cwd}\\Daily Stock Analysis\\Options\\{stock} Options '
                                      f'Data {dt.date.today()}.xlsx'):
                    url = f"{self.cwd}\\Daily Stock Analysis\\Options\\{stock} Options Data {self.today.date()}.xlsx"
                    wb = openpyxl.Workbook()
                    wb.save(url)

                    logging.debug(f"Missing excel document detected for the following stock: {stock}")
                    with sqlite3.connect(f'Databases/options.db') as db:
                        # resource for wildcard characters in SQL
                        # https://www.w3schools.com/sql/sql_wildcards.asp
                        table_names = db.execute(f"SELECT name from sqlite_master where name like '{stock}%'").fetchall()
                        for table in table_names:
                            if stock not in table[0]:
                                continue

                            book = openpyxl.load_workbook(url)
                            writer = pd.ExcelWriter(url, engine='openpyxl')
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            dataframe = pd.read_sql(f"select * from '{table[0]}'", db).set_index('strike')
                            dataframe.to_excel(writer, sheet_name=f'{table[0]}')
                            logger.debug(f"{table[0]} successfully outputted to Excel")

                            try:
                                sheet = book['Sheet']
                                book.remove(sheet)
                            except KeyError:
                                pass

                            writer.save()
                            writer.close()
                            book.save(url)
                            book.close()

                            ExcelFormatting(file_path=url).formatting()
            return

        with concurrent.futures.ThreadPoolExecutor(max_workers=len(self.stock_tickers)) as executor:
            for stock in outdated_options_chains:
                # when an updated options chain is not found
                executor.submit(self.initial_options, stock)

    def initial_options(self, stock):
        handle = ctypes.cdll.LoadLibrary(fr"{self.cwd}\Binaries\CallPricingDll.dll")
        #r"C:\Users\fabio\source\repos\CallPricingDll\CallPricingDll\x64\Release\CallPricingDll.dll")

        handle.CallPricing.argtypes = [ctypes.c_float, ctypes.c_float, ctypes.c_float, ctypes.c_float, ctypes.c_float]
        handle.CallPricing.restype = ctypes.c_double
        handle.PutPricing.argtypes = [ctypes.c_float, ctypes.c_float, ctypes.c_float, ctypes.c_float, ctypes.c_float]
        handle.PutPricing.restype = ctypes.c_double

        url = f"{self.cwd}\\Daily Stock Analysis\\Options\\{stock} Options Data {self.today.date()}.xlsx"

        wb = openpyxl.Workbook()
        wb.save(url)
        book = openpyxl.load_workbook(url)
        writer = pd.ExcelWriter(url, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        try:
            dividend = self.initial_data[stock][-1]['dividend']
            spot = self.quote_data[stock][-1]['current price']

            i = 0
            query = yahooquery.Ticker([stock], asynchronous=True)
            options = query.option_chain
            expiration_dates = list(options.index.unique(level=1))

            for date in expiration_dates:
                exp = date.to_pydatetime().date()
                # options expire at 3 o'clock CST
                exp_time = dt.datetime.combine(exp, dt.time(15, 0))
                time_diff = exp_time - self.today
                if time_diff.days < 0:
                    continue

                days_till_expiration = round(time_diff.total_seconds() / 86400, 2)
                if days_till_expiration > 100:
                    continue

                sofr = self.bond_bootstrapper.sofr()
                ois_rate = self.bond_bootstrapper.overnightindexedswaps(days_till_expiration, sofr)

                options_chain = options.loc[stock, date]
                call_table = options_chain.loc['calls']
                put_table = options_chain.loc['puts']

                call_table = call_table.assign(option_value=0.00).set_index('strike')
                put_table = put_table.assign(option_value=0.00).set_index('strike')
                self.option_value[stock].append({str(exp): {'overvalued_call_options': 0, 'undervalued_call_options': 0,
                                                            'overvalued_put_options': 0, 'undervalued_put_options': 0}})

                ois_rate -= dividend  # dividend should be factored in

                for index, row in call_table.iterrows():
                    # this means that there have been no trades over the past day
                    if row['change'] == 0:
                        continue

                    sigma = round(float(row['impliedVolatility']), 6)
                    strike = float(index)

                    option_price = handle.CallPricing(spot, strike, ois_rate, days_till_expiration, sigma)

                    call_table.at[index, 'option_value'] = round(option_price, 3)
                    spread = (row['bid'] + row['ask']) / 2
                    call_table.at[index, 'lastPrice'] = spread

                    if option_price > spread:
                        self.option_value[stock][i][str(exp)]['undervalued_call_options'] += 1
                    if option_price < spread:
                        self.option_value[stock][i][str(exp)]['overvalued_call_options'] += 1

                for index, row in put_table.iterrows():
                    # this means that there have been no trades over the past day
                    if row['change'] == 0:
                        continue

                    sigma = round(float(row['impliedVolatility']), 6)
                    strike = float(index)

                    option_price = handle.PutPricing(spot, strike, ois_rate, days_till_expiration, sigma)

                    put_table.at[index, 'option_value'] = round(option_price, 3)
                    spread = (row['bid'] + row['ask']) / 2
                    put_table.at[index, 'lastPrice'] = spread

                    if option_price > spread:
                        self.option_value[stock][i][str(exp)]['undervalued_put_options'] += 1
                    if option_price < spread:
                        self.option_value[stock][i][str(exp)]['overvalued_put_options'] += 1

                i += 1

                connection = sqlite3.connect(f'{self.cwd}\\Databases\\options.db')

                call_table.to_sql(name=f'{stock} Calls {exp}', con=connection, if_exists='replace')
                put_table.to_sql(name=f'{stock} Puts {exp}', con=connection, if_exists='replace')

                cursor = connection.cursor()
                ts = dt.datetime.now()

                # this ensures only one timestamp gets saved per stock
                table_no = self.stock_tickers.index(stock)
                cursor.execute("INSERT OR IGNORE INTO timestamp (stock, time, expiration, table_no) VALUES(?, ?, ?, ?)",
                               (stock, ts, exp_time, table_no))

                connection.commit()
                connection.close()

                # make sure we get nonzero values, since we can get some weirdness after hours and close to expirations
                if not unique_cols(call_table['option_value']):
                    call_table.to_excel(writer, sheet_name=f'{stock} Calls {exp}')
                    logger.debug(f"Calls for {stock} expiring {exp} successfully outputted to Excel")
                else:
                    logger.debug(f"Calls for {stock} expiring {exp} had no data to price options")

                if not unique_cols(put_table['option_value']):
                    put_table.to_excel(writer, sheet_name=f'{stock} Puts {exp}')
                    logger.debug(f"Puts for {stock} expiring {exp} successfully outputted to Excel")
                else:
                    logger.debug(f"Puts for {stock} expiring {exp} had no data to price options")

                try:
                    sheet = book['Sheet']
                    book.remove(sheet)
                except KeyError:
                    pass

        except Exception as e:
            logger.debug(f"Exception occurred: {e}")
        finally:
            writer.save()
            writer.close()
            book.save(url)
            book.close()

            ExcelFormatting(file_path=url).formatting()
