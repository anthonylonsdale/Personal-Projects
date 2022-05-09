import csv
import os
from pyexcelerate import Workbook
import openpyxl
import pandas as pd
import re
import glob


# this method of gathering all the data first then outputting to excel all in one go is much much faster
def export_to_excel(data):
    wb = Workbook()
    for number, data_list in enumerate(data):
        wb.new_sheet(wks_title_list[number], data=data_list)
    wb.save(excel_wb_title)
    return


# I want to make this dynamic so I don't have to keep hardcoding file names and changing it every time i do a new system
if __name__ == "__main__":
    # only change that is needed is to edit the target directory
    cwd = os.getcwd()
    target_directories = ['NoSpin']
    for directory in target_directories:
        systems = next(os.walk(directory))[1]
        if not systems:
            systems = ['']

        for system_size in systems:
            excel_wb_title = f"{system_size}_{directory} Python.xlsx"
            print(excel_wb_title)

            # this gets all the .heat files in the directory
            files = glob.glob(fr'{cwd}\\{directory}\\{system_size}\*.heat')

            data_list = []
            wks_title_list = []
            for file in files:
                data = []
                filename = file.split('\\')[-1]
                temp_string = re.search('.*_(.*k).*.heat', filename).group(1)
                wks_title_list.append(temp_string)
                print(temp_string)

                with open(file, 'r') as f:
                    reader = csv.reader(f)
                    for position, row in enumerate(reader):
                        if position > 2:
                            listrow = row[0]
                            listrow = list(listrow.split(" "))
                            listrow = list(filter(None, listrow))
                            listrow = map(float, listrow)
                            data.append(listrow)
                data_list.append(data)

            export_to_excel(data_list)
            data_list = None
            #################################################################################################
            xl = pd.ExcelFile(excel_wb_title)
            sheet_names = xl.sheet_names
            xl.close()

            wb = openpyxl.load_workbook(excel_wb_title)
            for sheet in sheet_names:
                ws = wb[sheet]
                count = 1

                temperature = int(sheet.split('k')[0])

                # this is for the formulas
                for i, cellObj in enumerate(ws['I'], 2):
                    if i > 21:
                        break

                    startingmaxrow = (ws.max_row - 20) + i
                    # cut out points at low temperatures because they are inaccurate as time goes on
                    #if int(temperature) == 50:
                    #    startingmaxrow -= 25000
                    #elif int(temperature) < 100:
                    #    startingmaxrow -= 20000

                    startingirow = 0 + i

                    cell = f'H{startingirow}'
                    icell = f'G{startingirow}'
                    formula = f"=AVERAGE(IF(MOD(_xlfn.ROW(D{startingirow}:D{startingmaxrow})-(A{startingirow}+1)," \
                              f"21)=0,D{startingirow}:D{startingmaxrow}))"

                    ws[cell] = formula
                    ws[icell] = count
                    ws.formula_attributes[cell] = {'t': 'array', 'ref': f"{cell}:{cell}"}

                    count += 1

                print(sheet)

                xvalues = openpyxl.chart.Reference(ws, min_col=7, min_row=2, max_col=7, max_row=21)
                yvalues = openpyxl.chart.Reference(ws, min_col=8, min_row=2, max_col=8, max_row=21)

                chart = openpyxl.chart.scatter_chart.ScatterChart()
                series = openpyxl.chart.Series(yvalues, xvalues)
                series.marker = openpyxl.chart.marker.Marker('circle')
                series.graphicalProperties.line.noFill = True
                chart.series.append(series)
                chart.title = "1-20"
                ws.add_chart(chart, "L2")

                xvalues1 = openpyxl.chart.Reference(ws, min_col=7, min_row=3, max_col=7, max_row=10)
                yvalues1 = openpyxl.chart.Reference(ws, min_col=8, min_row=3, max_col=8, max_row=10)

                chart = openpyxl.chart.scatter_chart.ScatterChart()
                series = openpyxl.chart.Series(yvalues1, xvalues1)
                series.marker = openpyxl.chart.marker.Marker('circle')
                series.graphicalProperties.line.noFill = True
                chart.series.append(series)
                chart.title = "2-9"
                series.trendline = openpyxl.chart.trendline.Trendline(dispEq=True, dispRSqr=True, trendlineType="linear")
                ws.add_chart(chart, "L18")

                xvalues1 = openpyxl.chart.Reference(ws, min_col=7, min_row=12, max_col=7, max_row=20)
                yvalues1 = openpyxl.chart.Reference(ws, min_col=8, min_row=12, max_col=8, max_row=20)

                chart = openpyxl.chart.scatter_chart.ScatterChart()
                series = openpyxl.chart.Series(yvalues1, xvalues1)
                series.marker = openpyxl.chart.marker.Marker('circle')
                series.graphicalProperties.line.noFill = True
                chart.series.append(series)
                chart.title = "11-19"
                series.trendline = openpyxl.chart.trendline.Trendline(dispEq=True, dispRSqr=True, trendlineType="linear")
                ws.add_chart(chart, "L34")

                # slopes of 2-9 and 11-19
                y_point_1 = 'H3'
                x_point_1 = 'G3'
                y_point_2 = 'H10'
                x_point_2 = 'G10'
                y_point_3 = 'H13'
                x_point_3 = 'G13'
                y_point_4 = 'H20'
                x_point_4 = 'G20'

                if int(temperature) <= 100:
                    y_point_1 = 'H4'
                    x_point_1 = 'G4'
                    y_point_2 = 'H11'
                    x_point_2 = 'G11'

                ws['G24'] = f"=LINEST({y_point_1}:{y_point_2}, {x_point_1}:{x_point_2}, TRUE, TRUE)"
                ws.formula_attributes['G24'] = {'t': 'array', 'ref': "G24:H28"}

                ws['I24'] = f"=LINEST({y_point_3}:{y_point_4}, {x_point_3}:{x_point_4}, TRUE, TRUE)"
                ws.formula_attributes['I24'] = {'t': 'array', 'ref': "I24:J28"}

            wb.save(excel_wb_title)
            wb.close()
            wb = None
            print("wb saved, program free to terminate")
