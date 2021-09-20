import csv
import pandas as pd
import xlsxwriter
import openpyxl
import os

if __name__ == "__main__":
    systems = ['NoSpin', 'Ferro', 'Anti-Ferro']

    for j in range(len(systems)):
        title = "{} Heat Data Python.xlsx".format(systems[j])

        if os.path.isfile(r"{}".format(title)):
            os.remove(r"{}".format(title))

        wb = openpyxl.Workbook()
        wb.save(title)

        temperature = 50
        for i in range(1, 21):
            main_list = []
            temp = int(temperature) * i
            temp_string = str(temp)
            row = 0
            print(temp_string)

            book = openpyxl.load_workbook(title)
            wks_title = temp_string + str('k')
            writer = pd.ExcelWriter(title, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

            workbook = xlsxwriter.Workbook(title)
            worksheet = workbook.add_worksheet(wks_title)

            if j == 0:
                directory = r"C:\Users\fabio\PycharmProjects\Old Projects\UO2 Potential" \
                            r"\{}\{}_profile_{}k.heat".format(systems[j], systems[j], temp_string)
            else:
                directory = r"C:\Users\fabio\PycharmProjects\Old Projects\UO2 Potential" \
                            r"\{}\{}_{}k_profile.heat".format(systems[j], systems[j], temp_string)

            with open(directory, 'r') as f:
                reader = csv.reader(f)
                for position, row in enumerate(reader):
                    if j == 0:
                        if i == 1:
                            if position >= 80917:
                                listrow = row[0]
                                listrow = list(listrow.split(" "))
                                listrow = list(filter(None, listrow))
                                listrow = map(float, listrow)
                                main_list.append(listrow)
                        if position >= 102484:
                            listrow = row[0]
                            listrow = list(listrow.split(" "))
                            listrow = list(filter(None, listrow))
                            listrow = map(float, listrow)
                            main_list.append(listrow)

                    elif j == 1:
                        if position >= 9223:
                            listrow = row[0]
                            listrow = list(listrow.split(" "))
                            listrow = list(filter(None, listrow))
                            listrow = map(float, listrow)
                            main_list.append(listrow)
                    else:
                        if position >= 9244:
                            listrow = row[0]
                            listrow = list(listrow.split(" "))
                            listrow = list(filter(None, listrow))
                            listrow = map(float, listrow)
                            main_list.append(listrow)

            df = pd.DataFrame(main_list)
            df.to_excel(writer, sheet_name=wks_title)
            writer.save()

            try:
                sheet = book['Sheet']
                book.remove(sheet)
                book.save(title)
            except KeyError:
                pass
