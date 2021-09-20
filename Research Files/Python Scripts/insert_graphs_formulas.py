import openpyxl.chart


if __name__ == '__main__':
    systems = ['NoSpin', 'Ferro', 'Anti-Ferro']

    for j in range(len(systems)):

        title = "{} Heat Data Python.xlsx".format(systems[j])
        temperature = 50

        wb = openpyxl.load_workbook(title)

        for i in range(1, 21):
            temp = int(temperature) * i
            temp_string = str(temp)
            wks_title = temp_string + str('k')

            print(wb[wks_title])

            ws = wb.get_sheet_by_name(wks_title)
            count = 1

            for i, cellObj in enumerate(ws['I'], 2):
                if i > 21:
                    break

                if j == 0:
                    startingmaxrow = 6300 + i
                elif j == 1:
                    startingmaxrow = 1260 + i
                else:
                    startingmaxrow = 1260 + i
                startingirow = 0 + i

                cell = 'H{}'.format(startingirow)
                icell = 'G{}'.format(startingirow)
                formula = f"=AVERAGE(IF(MOD(_xlfn.ROW(E{startingirow}:E{startingmaxrow})-(B{startingirow}+1)," \
                          f"21)=0,E{startingirow}:E{startingmaxrow}))"

                ws[cell] = formula
                ws[icell] = count
                ws.formula_attributes[cell] = {'t': 'array', 'ref': f"{cell}:{cell}"}
                count += 1

            xvalues = openpyxl.chart.Reference(ws, min_col=7, min_row=2, max_col=7, max_row=21)
            yvalues = openpyxl.chart.Reference(ws, min_col=8, min_row=2, max_col=8, max_row=21)

            chart = openpyxl.chart.scatter_chart.ScatterChart()
            series = openpyxl.chart.Series(yvalues, xvalues)
            series.marker = openpyxl.chart.marker.Marker('circle')
            series.graphicalProperties.line.noFill = True
            chart.series.append(series)
            chart.title = "1-20"
            ws.add_chart(chart, "J3")

            xvalues1 = openpyxl.chart.Reference(ws, min_col=7, min_row=3, max_col=7, max_row=10)
            yvalues1 = openpyxl.chart.Reference(ws, min_col=8, min_row=3, max_col=8, max_row=10)

            chart = openpyxl.chart.scatter_chart.ScatterChart()
            series = openpyxl.chart.Series(yvalues1, xvalues1)
            series.marker = openpyxl.chart.marker.Marker('circle')
            series.graphicalProperties.line.noFill = True
            chart.series.append(series)
            chart.title = "2-9"
            series.trendline = openpyxl.chart.trendline.Trendline(dispEq=True, dispRSqr=True, trendlineType="linear")
            ws.add_chart(chart, "J19")

            xvalues1 = openpyxl.chart.Reference(ws, min_col=7, min_row=12, max_col=7, max_row=20)
            yvalues1 = openpyxl.chart.Reference(ws, min_col=8, min_row=12, max_col=8, max_row=20)

            chart = openpyxl.chart.scatter_chart.ScatterChart()
            series = openpyxl.chart.Series(yvalues1, xvalues1)
            series.marker = openpyxl.chart.marker.Marker('circle')
            series.graphicalProperties.line.noFill = True
            chart.series.append(series)
            chart.title = "11-19"
            series.trendline = openpyxl.chart.trendline.Trendline(dispEq=True, dispRSqr=True, trendlineType="linear")
            ws.add_chart(chart, "J35")

        wb.save(title)
