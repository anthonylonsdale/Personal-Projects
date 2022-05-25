import csv
import os
from pyexcelerate import Workbook
import openpyxl
import pandas as pd
import re
import glob


def export_to_excel(data):
    wb = Workbook()
    for number, data_list in enumerate(data):
        wb.new_sheet(wks_title_list[number], data=data_list)
    wb.save(excel_wb_title)
    return


if __name__ == '__main__':
    cwd = os.getcwd()
    files = glob.glob(fr'{cwd}\\10-1000 chunks\*.lammpstrj')
    excel_wb_title = f"anti_ferro_magnetization.xlsx"

    data_list = []
    wks_title_list = []
    for file in files:
        data = []
        temp_string = re.search('.*_(.*k).*.lammpstrj', file).group(1)
        wks_title_list.append(temp_string)
        print(temp_string)

        with open(file, 'r') as f:
            reader = csv.reader(f)
            for position, row in enumerate(reader):
                if position > 2:
                    listrow = row[0]
                    listrow = list(listrow.split(" "))
                    listrow = list(filter(None, listrow))
                    listrow = map(float, listrow)
                    data.append(listrow)
        data_list.append(data)

    export_to_excel(data_list)
    data_list = None

    xl = pd.ExcelFile(excel_wb_title)
    sheet_names = xl.sheet_names
    xl.close()

    wb = openpyxl.load_workbook(excel_wb_title)
    for sheet in sheet_names:
        ws = wb[sheet]
        count = 1

        temperature = int(sheet.split('k')[0])

        # this is for the formulas
        for i, cellObj in enumerate(ws['I'], 2):
            if i > 21:
                break

            startingmaxrow = (ws.max_row - 21) + i
            startingirow = 0 + i

            cell = f'I{startingirow}'
            icell = f'H{startingirow}'

            formula = f"=SQRT(D{startingmaxrow}^2+E{startingmaxrow}^2+F{startingmaxrow}^2)"

            ws[cell] = formula
            ws[icell] = count
            ws.formula_attributes[cell] = {'t': 'array', 'ref': f"{cell}:{cell}"}

            count += 1

        ws["I23"] = "=AVERAGE(I2:I21)"

    wb.create_sheet('antiferromagnetic values')
    ws = wb['antiferromagnetic values']
    for i in range(1, 101):
        temp = i * 10
        sheet_name = f'{temp}k'
        cell = f'A{i}'
        icell = f'B{i}'
        ws[cell] = temp
        ws[icell] = f"='{sheet_name}'!I23"

    wb.save(excel_wb_title)
    wb.close()
    wb = None

    print("wb saved, program free to terminate")
