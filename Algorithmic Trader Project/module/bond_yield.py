from requests import get
from bs4 import BeautifulSoup
import pandas as pd
from pandas.tseries.offsets import BDay
import openpyxl
import datetime as dt
"""
The methodology will be expanded upon later but essentially
we can take a look at changing bond prices as an indicator
of market volatility, if bond prices increase then that
signals more investors are purchasing bonds, causing
yields to decrease, and the opposite is true.
"""

class treasuryYields:
    def __init__(self):
        self.bond_yields = {}
        self.todays_date = dt.datetime.today() - BDay(1)
        self.url = f"C:/Users/fabio/PycharmProjects/AlgoTrader/ALGO/Daily Stock Analysis/Bonds/T-Bonds Data " \
                   f"{self.todays_date.date()}.xlsx"
        wb = openpyxl.Workbook()
        wb.save(self.url)
        self.r = get("https://www.treasury.gov/resource-center/data-chart-center/interest-rates/Pages/Text"
                     "View.aspx?data=yield")

    def treasury_bond_yields(self):
        soup = BeautifulSoup(self.r.text, 'lxml')
        bond_list = [entry.text for entry in soup.find_all('td', {'class': 'text_view_data'})]
        date = None
        for i in range(len(bond_list)):
            if i % 13 == 0:
                date = bond_list[i]
                self.bond_yields[date] = []
                continue
            self.bond_yields[date].append(bond_list[i])

        bond_df = pd.DataFrame.from_dict(self.bond_yields).T
        bond_df = bond_df.rename(columns={0: '1 Mo', 1: '2 Mo', 2: '3 Mo', 3: '6 Mo', 4: '1 Yr', 5: '2 Yr',
                                          6: '3 Yr', 7: '5 Yr', 8: '7 Yr', 9: '10 Yr', 10: '20 Yr', 11: '30 Yr'})
        string_date = date.replace("/", '-')
        wb = openpyxl.Workbook()
        wb.save(self.url)
        book = openpyxl.load_workbook(self.url)
        writer = pd.ExcelWriter(self.url, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        bond_df.to_excel(writer, sheet_name=f"Treasury Bonds for {string_date}")
        try:
            sheet = book['Sheet']
            book.remove(sheet)
        except KeyError:
            pass
        book.save(self.url)

        onemonthyield = bond_df.iloc[13, 0]
        twomonthyield = bond_df.iloc[13, 1]
        threemonthyield = bond_df.iloc[13, 2]
        sixmonthyield = bond_df.iloc[13, 3]
        oneyryield = bond_df.iloc[13, 4]
        twoyryield = bond_df.iloc[13, 5]
        return (onemonthyield, twomonthyield, threemonthyield,
                sixmonthyield, oneyryield, twoyryield)


if __name__ == "__main__":
    bond_object = treasuryYields()
    bond_object.treasury_bond_yields()
