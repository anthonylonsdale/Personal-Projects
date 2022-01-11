import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
import os
import datetime as dt
from dateutil.parser import parse
import logging

from ALGO.excel_formatting_module import ExcelFormatting

logger = logging.getLogger(__name__)
# added logging statements

"""
Replace the government bond yield curve (which is quite wrong to use) and either use the LIBOR
rates or the OIS (Overnight Index Swap) rates
https://www-2.rotman.utoronto.ca/~hull/DownloadablePublications/LIBORvsOIS.pdf
Traditionally practitioners have used LIBOR and LIBOR-swap rates as proxies for risk-free rates
when valuing derivatives. This practice has been called into question by the credit crisis that
started in 2007. Many banks now consider that overnight indexed swap (OIS) rates should be
used as the risk-free rate when collateralized portfolios are valued and that LIBOR should be
used for this purpose when portfolios are not collateralized. This paper examines this practice
and concludes that OIS rates should be used in all situations.
"""


class bondYields:
    def __init__(self):
        logger.debug("Initialization of bond gathering")
        self.bond_yields = {}
        self.todays_date = dt.datetime.today()
        self.cwd = os.getcwd()

    # preferred rate for pricing derivatives
    # the rate at which you can borrow/lend cash is the rate you need to consider
    def overnightindexedswaps(self, days_to_expiry, overnight_rate):
        # the theory behind this is to calculate the spread between a compounded daily index rate (the overnight rate)
        # and a fixed index rate. this is valid up to a two week term

        # first step is to multiply the overnight rate for the period in which the swap applies
        # the overnight lending rate is typically the federal funds rate or the SOFR (secured overnight financing rate)
        floating_rate_leg = ((overnight_rate * days_to_expiry) / 360) + 1


        # fixed rate is the LIBOR rate as it is the least risky rate available
        # this also should pretty much always be larger
        fixed_rate_leg = ((overnight_rate / 360) + 1) ** days_to_expiry

        ois_rate = (fixed_rate_leg - floating_rate_leg) * 100
        logging.debug(f"OIS Rate: {ois_rate}% given overnight rate of {overnight_rate}% "
                      f"with {days_to_expiry} days to expiration")
        return round(ois_rate, 6)

    # secured overnight financing rate is the rate at which banks lend to each other overnight, fixed leg of OIS
    def sofr(self):
        with requests.get("https://fred.stlouisfed.org/series/SOFR", stream=True) as r:
            soup = BeautifulSoup(r.text, 'lxml')
            sofr_list = [entry.text for entry in soup.find_all('span', {'class': 'series-meta-observation-value'})]
            sofr = float(sofr_list[-1])
        logging.debug(f"SOFR rate: {sofr}%")
        return sofr

    # LIBOR rates as used as the floating leg in the OIS calculation
    def LIBOR_yields(self):
        with requests.get("https://www.global-rates.com/en/interest-rates/libor/american-dollar/american-dollar.aspx",
                          stream=True) as r:
            soup = BeautifulSoup(r.text, 'lxml')
            libor_list = [entry.text for entry in soup.find_all('td', {'align': 'center'})]

        libor_2dlist = []
        row = []
        date = libor_list[1]
        for i in range(1, 81):
            element = libor_list[i]
            if i <= 5:
                element = dt.datetime.strptime(element, '%m-%d-%Y').strftime('%Y-%m-%d')
            row.append(element)
            if i % 5 == 0:
                libor_2dlist.append(row)
                row = []

        libor_df = pd.DataFrame.from_records(libor_2dlist)
        libor_refitted_df = libor_df.set_axis(libor_2dlist[0], axis='columns')
        libor_refitted_df = libor_refitted_df.iloc[1:, :]
        terms = ['USD-LIBOR-OVERNIGHT', 'USD-LIBOR-1-WEEK', 'USD-LIBOR-2-WEEKS', 'USD-LIBOR-1-MONTH',
                 'USD-LIBOR-2-MONTHS', 'USD-LIBOR-3-MONTHS', 'USD-LIBOR-4-MONTHS', 'USD-LIBOR-5-MONTHS',
                 'USD-LIBOR-6-MONTHS', 'USD-LIBOR-7-MONTHS', 'USD-LIBOR-8-MONTHS', 'USD-LIBOR-9-MONTHS',
                 'USD-LIBOR-10-MONTHS', 'USD-LIBOR-11-MONTHS', 'USD-LIBOR-12-MONTHS']
        libor_refitted_df.index = terms

        date = pd.to_datetime(date).date()
        url = self.cwd + fr"\Daily Stock Analysis\Bonds\LIBOR Yields (last updated {date}).xlsx"
        wb = openpyxl.Workbook()
        wb.save(url)
        book = openpyxl.load_workbook(url)
        writer = pd.ExcelWriter(url, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        libor_refitted_df.to_excel(writer, sheet_name=f"LIBOR yields for {date}")
        try:
            sheet = book['Sheet']
            book.remove(sheet)
        except KeyError:
            pass
        writer.save()
        writer.close()
        book.save(url)
        book.close()

        ExcelFormatting(file_path=url).formatting()

        for i in range(4):
            try:
                overnight_rate = float(libor_refitted_df.iloc[0, i].split()[0])
                one_week_rate = float(libor_refitted_df.iloc[1, i].split()[0])
                one_month_rate = float(libor_refitted_df.iloc[3, i].split()[0])
                two_month_rate = float(libor_refitted_df.iloc[4, i].split()[0])
                three_month_rate = float(libor_refitted_df.iloc[5, i].split()[0])
                six_month_rate = float(libor_refitted_df.iloc[8, i].split()[0])
                twelve_month_rate = float(libor_refitted_df.iloc[14, i].split()[0])
                libor_yields = (round(overnight_rate, 6), round(one_week_rate, 6), round(one_month_rate, 6),
                                round(two_month_rate, 6), round(three_month_rate, 6), round(six_month_rate, 6),
                                round(twelve_month_rate, 6))
                logging.debug(libor_yields)
                return libor_yields
            except ValueError:
                continue

    def treasury_bond_yields(self):
        with requests.get("https://www.treasury.gov/resource-center/data-chart-center/interest-"
                          "rates/Pages/TextView.aspx?data=yield", stream=True) as r:
            soup = BeautifulSoup(r.text, 'lxml')
            bond_list = [entry.text for entry in soup.find_all('td', {'class': 'text_view_data'})]

        date = None
        for i in range(len(bond_list)):
            if i % 13 == 0:
                date = parse(bond_list[i].replace("/", '-')).date()
                self.bond_yields[date] = []
                continue
            self.bond_yields[date].append(bond_list[i])

        bond_df = pd.DataFrame.from_dict(self.bond_yields).T
        bond_df = bond_df.rename(columns={0: '1 Mo', 1: '2 Mo', 2: '3 Mo', 3: '6 Mo', 4: '1 Yr', 5: '2 Yr',
                                          6: '3 Yr', 7: '5 Yr', 8: '7 Yr', 9: '10 Yr', 10: '20 Yr', 11: '30 Yr'})

        url = self.cwd + f"\\Daily Stock Analysis\\Bonds\\US T-Bond Yields (last updated {date}).xlsx"
        wb = openpyxl.Workbook()
        wb.save(url)
        book = openpyxl.load_workbook(url)
        writer = pd.ExcelWriter(url, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        bond_df.to_excel(writer, sheet_name=f"Treasury Bonds for {date}")
        try:
            sheet = book['Sheet']
            book.remove(sheet)
        except KeyError:
            pass
        writer.save()
        writer.close()
        book.save(url)
        book.close()

        ExcelFormatting(file_path=url).formatting()
        logging.debug("Treasury bond excel sheet saved successfully")

        onemonthyield = float(bond_df.iloc[-1, 0]) / 100
        twomonthyield = float(bond_df.iloc[-1, 1]) / 100
        threemonthyield = float(bond_df.iloc[-1, 2]) / 100
        sixmonthyield = float(bond_df.iloc[-1, 3]) / 100
        oneyryield = float(bond_df.iloc[-1, 4]) / 100
        twoyryield = float(bond_df.iloc[-1, 5]) / 100

        yield_tuple = (round(onemonthyield, 6), round(twomonthyield, 6), round(threemonthyield, 6),
                       round(sixmonthyield, 6), round(oneyryield, 6), round(twoyryield, 6))
        logging.debug(yield_tuple)

        return yield_tuple
