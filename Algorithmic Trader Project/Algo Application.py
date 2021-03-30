# updated to full functionality

import yfinance as yf
import pandas as pd
import shutil
import glob
import requests
from win10toast import ToastNotifier
from requests import get
import websocket
from json import loads
import threading as th
import alpaca_trade_api as trade_api
import datetime as dt
from clr import AddReference
import csv
from pandas import DataFrame
import openpyxl
import win32com.client as win32
import os
from bs4 import BeautifulSoup
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import numpy as np
import time


def get_tickers():
    tickers = []
    table = pd.read_html('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies')
    df = table[0]
    sandp500tickers = df['Symbol'].to_list()
    print(sandp500tickers)

    # these headers and params are subject to change in the event NASDAQ changes its API
    headers = {'authority': 'api.nasdaq.com', 'accept': 'application/json, text/plain, */*',
               'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko)'
                             ' Chrome/87.0.4280.141 Safari/537.36',
               'origin': 'https://www.nasdaq.com', 'sec-fetch-site': 'same-site', 'sec-fetch-mode': 'cors',
               'sec-fetch-dest': 'empty', 'referer': 'https://www.nasdaq.com/', 'accept-language': 'en-US,en;q=0.9', }
    params = (('tableonly', 'true'), ('limit', '25'), ('offset', '0'), ('download', 'true'),)

    r = requests.get('https://api.nasdaq.com/api/screener/stocks', headers=headers, params=params)
    data = r.json()['data']
    df = pd.DataFrame(data['rows'], columns=data['headers'])
    df = df[df.country == 'United States']
    df['marketCap'] = pd.to_numeric(df['marketCap'])
    df['volume'] = pd.to_numeric(df['volume'])
    df = df[df.marketCap > 500000000]
    df = df[df.marketCap != 0 & df.marketCap.notnull()]
    df = df[df.volume > 2340000]

    for index, row in df.iterrows():
        for i in range(len(sandp500tickers)):
            stock = row['symbol']
            if stock == sandp500tickers[i]:
                try:
                    asset = api.get_asset(stock)
                    print(asset)
                    if asset.tradable and asset.easy_to_borrow and asset.marginable and asset.shortable:
                        tickers.append(stock)
                except Exception as error:
                    print(error)
                    pass

    book = openpyxl.load_workbook('Ticker Selections.xlsx')
    writer = pd.ExcelWriter('Ticker Selections.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer)
    try:
        sheet = book['Sheet']
        book.remove(sheet)
        book.save('Ticker Selections.xlsx')
    except KeyError:
        pass
    writer.save()
    return tickers


def api_calls(tickers):
    yfinance_api_calls = 0
    stock_failures = 0
    stocks_not_imported = 0
    i = 0
    while (i < len(tickers)) and (yfinance_api_calls < 1800):
        try:
            stock = tickers[i]
            temp = yf.Ticker(str(stock))
            his_data = temp.history(period="max")
            his_data.to_csv(r"Daily Stock Analysis\Stocks\\" + stock + ".csv")
            time.sleep(2)
            yfinance_api_calls += 1
            stock_failures = 0
            i += 1
            print('Number of stocks gathered:', i)
        except ValueError:
            print("Error with gathering Yahoo Finance data for {}".format(str(tickers[i])))
            if stock_failures > 5:
                i += 1
                stocks_not_imported += 1
            yfinance_api_calls += 1
            stock_failures += 1
    print("Data successfully imported for {} stocks".format(i - stocks_not_imported))


def obv_score_array():
    list_files = (glob.glob(r"Daily Stock Analysis\Stocks\\*.csv"))
    obv_score_array = []
    interval = 0
    while interval < len(list_files):
        data = pd.read_csv(list_files[interval]).tail(7)
        pos_move = []
        neg_move = []
        obv_value = 0
        count = 0
        while count < 7:
            if data.iloc[count, 1] < data.iloc[count, 4]:
                pos_move.append(count)
            elif data.iloc[count, 1] > data.iloc[count, 4]:
                neg_move.append(count)
            count += 1
        for i in pos_move:
            obv_value = round(obv_value + (data.iloc[i, 5] / data.iloc[i, 1]))
        for i in neg_move:
            obv_value = round(obv_value - (data.iloc[i, 5] / data.iloc[i, 1]))
        stock_name = ((os.path.basename(list_files[interval])).split(".csv")[0])
        obv_score_array.append([stock_name, obv_value])
        interval += 1

    obv_dataframe = pd.DataFrame(obv_score_array, columns=['Stock', 'OBV_Value'])
    obv_dataframe["Stocks_Ranked"] = obv_dataframe["OBV_Value"].rank(ascending=False)
    obv_dataframe.sort_values("OBV_Value", inplace=True, ascending=False)
    print(obv_dataframe)
    obv_dataframe.to_csv("Daily Stock Analysis/OBV_Ranked.csv", index=False)


"""

END OF ANALYSIS FUNCTIONS
START OF ALGO MAIN FUNCTIONS

"""


def on_message(ws, message):
    if message == '{"type":"ping"}':
        return
    data = loads(message)
    stock_fundamentals = data['data'][0]
    time_integer = stock_fundamentals['t']
    timestamp = dt.datetime.fromtimestamp(time_integer / 1e3)
    timestamp = str("{}.{:03d}".format(timestamp.strftime('%Y-%m-%d %H:%M:%S'), timestamp.microsecond // 1000))
    stock_fundamentals['t'] = timestamp
    s = stock_fundamentals['s']
    stock_fundamentals['time'] = stock_fundamentals.pop('t')
    stock_fundamentals['price'] = stock_fundamentals.pop('p')
    stock_fundamentals['stock'] = stock_fundamentals.pop('s')
    stock_fundamentals['volume'] = stock_fundamentals.pop('v')
    print(stock_fundamentals)
    trade_data[s].append(stock_fundamentals)


def on_error(ws, error):
    print(error)


def on_close(ws):
    print("### closed ###")


def on_open(ws):
    for stock_ticker in stock_tickers:
        custom_call = str('{"type":"subscribe","symbol":"') + stock_ticker + str('"}')
        print(custom_call)
        ws.send(custom_call)


def stock_data_engine():
    AddReference(r"C:\Users\fabio\source\repos\Webscraper Class Library\Webscraper Class Library\bin\Debug\Web"
                 r"scraper Class Library.dll")
    import CSharpwebscraper
    scraper = CSharpwebscraper.Webscraper()
    stock_info = scraper.Scraper(stock_tickers)
    print(stock_info)
    quote = {}
    for i in range(len(stock_info)):
        if (i % 6) == 0:
            quote['current price'] = float(stock_info[i].replace(",", ""))
        if (i % 6) == 1:
            quote['open price'] = float(stock_info[i].replace(",", ""))
        if (i % 6) == 2:
            quote['previous close'] = float(stock_info[i].replace(",", ""))
        if (i % 6) == 3:
            quote['indicator'] = str(stock_info[i])
        if (i % 6) == 4:
            delimiter1 = '('
            delimiter2 = ')'
            div = str(stock_info[i])
            dividend = div[div.find(delimiter1) + 1: div.find(delimiter2)]
            if dividend == 'N/A':
                dividend = 0
            else:
                div_string = dividend.split("%")[0]
                dividend = float(div_string) / 100
            quote['dividend'] = dividend
        if (i % 6) == 5:
            st = str(stock_info[i])
            quote['stock'] = st
            quote['time'] = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            quote_data[st].append(quote)
            quote = {}
            df_quote = DataFrame(quote_data[st])
            df_quote.set_index('time', inplace=True)
            print(df_quote)


def tech_indicators():
    resolution = ['1', '5', '15', '30', '60', 'D', 'W', 'M']
    for st in stock_tickers:
        i = 0
        while True:
            try:
                timestamp_tech = time.strftime("%Y-%m-%d %H:%M:%S")
                tech_url = 'https://finnhub.io/api/v1/scan/technical-indicator?symbol=' + st + \
                           '&resolution=' + resolution[i] + '&token=bsq43v0fkcbdt6un6ivg'
                r = get(tech_url)
                ti = r.json()
                technical = ti['technicalAnalysis']['count']
                technical['signal'] = ti['technicalAnalysis']['signal']
                technical['adx'] = ti['trend']['adx']
                technical['trending'] = ti['trend']['trending']
                technical['time'] = timestamp_tech
                ti_data[st].append(technical)
                df_tech = DataFrame(ti_data[st])
                df_tech.set_index('signal', inplace=True)
                print(df_tech)
            except KeyError:
                if i == 7:
                    break
                i += 1
            break


def main_data_engine():
    try:
        t1 = th.Thread(target=socket.run_forever)
        t1.daemon = True
        t1.start()
        t2 = th.Thread(target=stock_data_engine)
        t2.daemon = True
        t2.start()
        t3 = th.Thread(target=tech_indicators)
        t3.daemon = True
        t3.start()
        time.sleep(15)
        socket.keep_running = False
        socket.close()
        t1.join()
        t2.join()
        t3.join()
    except Exception as error:
        print(error)
        socket.close()


def ticker_operations():
    for element in stock_tickers:
        timestamp = dt.datetime.strptime(quote_data[element][-1]['time'][11::], '%H:%M:%S').time()
        indicator = quote_data[element][-1]['indicator']
        if len(ti_data) > 0:
            buy_votes = int(ti_data[element][-1]['buy'])
            sell_votes = int(ti_data[element][-1]['sell'])
            neutral_votes = int(ti_data[element][-1]['neutral'])
            indicator_votes[element]['Bullish Votes'] += buy_votes
            indicator_votes[element]['Bearish Votes'] += sell_votes
            indicator_votes[element]['Neutral Votes'] += neutral_votes

        if indicator == 'Bullish':
            indicator_votes[element]['Bullish Votes'] += 1
        if indicator == 'Bearish':
            indicator_votes[element]['Bearish Votes'] += 1

        if indicator_votes[element]['Bullish Votes'] > indicator_votes[element]['Bearish Votes'] and \
                indicator_votes[element]['Bullish Votes'] > indicator_votes[element]['Neutral Votes']:
            stock_buylist[element].append('Very Bullish at: ' + str(timestamp))
        elif indicator_votes[element]['Bullish Votes'] > indicator_votes[element]['Bearish Votes']:
            stock_buylist[element].append('Bullish at: ' + str(timestamp))

        if indicator_votes[element]['Bearish Votes'] > indicator_votes[element]['Bullish Votes'] and \
                indicator_votes[element]['Bearish Votes'] > indicator_votes[element]['Neutral Votes']:
            stock_shortlist[element].append('Very Bearish at: ' + str(timestamp))
        elif indicator_votes[element]['Bearish Votes'] > indicator_votes[element]['Bullish Votes']:
            stock_shortlist[element].append('Bearish at: ' + str(timestamp))

    print('Stocks of interest:', stock_tickers)
    print('Buy Side Stocklist:', stock_buylist)
    print('Sell Side Stocklist:', stock_shortlist)
    print('------------------------------------------------------------------------')


def volume_operations():
    for element in stock_tickers:
        volume_past_30sec = 0
        volume_past_min = 0
        volume_past_2min = 0
        volume_past_5min = 0
        stock_price = 0
        stock_list_length[element] = 0
        length = []
        stock_prices[element] = []
        for pos, ite in enumerate(trade_data[element]):
            if pos >= int(len(trade_data[element]) - 25):
                stock_price += ite['price']
                length.append(ite)
                stock_list_length[element] = int(len(length))
            ##############################################################################
            tradetime = dt.datetime.strptime(ite['time'], "%Y-%m-%d %H:%M:%S.%f")
            if tradetime > dt.datetime.now() - dt.timedelta(seconds=30):
                volume_past_30sec += int(ite['volume'])
            if tradetime > dt.datetime.now() - dt.timedelta(seconds=60):
                volume_past_min += int(ite['volume'])
            if tradetime > dt.datetime.now() - dt.timedelta(seconds=120):
                volume_past_2min += int(ite['volume'])
            if tradetime > dt.datetime.now() - dt.timedelta(seconds=300):
                volume_past_5min += int(ite['volume'])
        ##############################################################################
        try:
            stock_prices[element] = float("{:.3f}".format(stock_price / stock_list_length[element]))
        except ZeroDivisionError:
            stock_prices[element] = quote_data[element][-1]['current price']
            pass
        volume_terms[element] = volume_past_30sec, volume_past_min, volume_past_2min, volume_past_5min
        ##############################################################################
    print('volume by stock ordered 30sec, 1min, 2min and 5min:', volume_terms)
    print('stock prices:', stock_prices)


# might need rework and additional analysis
def analysis_operations():
    # this block is only to see if the stock has had an increase or decrease in short term price respectively
    # if the following conditions are met, then this represents a good trade opportunity
    # this portion works
    for sto in stock_buylist:
        for pos, i in enumerate(quote_data[sto]):
            if dt.datetime.strptime(quote_data[sto][pos]['time'], "%Y-%m-%d %H:%M:%S") > \
                    (dt.datetime.now() - dt.timedelta(minutes=5)):
                if pos == (len(quote_data[sto]) - 1) and quote_data[sto][pos]['current price'] < \
                        stock_prices[sto]:
                    stock_price_movement[sto] = 'short-term increase in price'
    for sto in stock_shortlist:
        for pos, i in enumerate(quote_data[sto]):
            if dt.datetime.strptime(quote_data[sto][pos]['time'], "%Y-%m-%d %H:%M:%S") > \
                    (dt.datetime.now() - dt.timedelta(minutes=5)):
                # if current price is lower than quote price
                if pos == (len(quote_data[sto]) - 1) and quote_data[sto][pos]['current price'] > \
                        stock_prices[sto]:
                    stock_price_movement[sto] = 'short-term decrease in price'
    print(stock_price_movement)
    ##############################################################################
    for element in stock_price_movement:
        if 'increase' in stock_price_movement[element]:
            if len(stock_buylist[element]) > 2:
                if quote_data[element][-1]['current price'] > quote_data[element][-2]['current price'] > \
                        quote_data[element][-3]['current price']:
                    if 'Very Bullish' in stock_buylist[element][-1] and 'Very Bullish' in stock_buylist[element][-2]:
                        strong_buy.append(element)
                elif quote_data[element][-1]['current price'] > quote_data[element][-2]['current price']:
                    if 'Bullish' in stock_buylist[element][-1] and 'Bullish' in stock_buylist[element][-2]:
                        buy.append(element)
                else:
                    weak_buy.append(element)
        if 'decrease' in stock_price_movement[element]:
            if len(stock_shortlist[element]) > 2:
                if quote_data[element][-1]['current price'] < quote_data[element][-2]['current price'] < \
                        quote_data[element][-3]['current price']:
                    if 'Very Bearish' in stock_shortlist[element][-1] and 'Very Bearish' in element:
                        strong_sell.append(element)
                elif quote_data[element][-1]['current price'] < quote_data[element][-2]['current price']:
                    if 'Bearish' in stock_shortlist[element][-1] and 'Bearish' in stock_shortlist[element][-2]:
                        sell.append(element)
                else:
                    weak_sell.append(element)
        ##############################################################################
    if len(strong_buy) > 0:
        print('Stock Strong Buy List:', strong_buy)
    if len(buy) > 0:
        print('Stock Buy List:', buy)
    if len(weak_buy) > 0:
        print('Stock Weak Buy List:', weak_buy)
    if len(strong_sell) > 0:
        print('Stock Strong Sell List:', strong_sell)
    if len(sell) > 0:
        print('Stock Sell List:', sell)
    if len(weak_sell) > 0:
        print('Stock Weak Sell List:', weak_sell)


def trade_execution_operations():
    global strong_buy, buy, weak_buy, strong_sell, sell, weak_sell, current_stock_position
    print(api.list_orders())
    block_purchase = []

    if len(api.list_positions()) > 0:
        for stockticker in stock_tickers:
            try:
                stock_position = api.get_position(stockticker)
                print(stock_position)
                position_value = getattr(stock_position, "market_value")
                position_value = abs(float(position_value))
                if position_value >= (0.10 * account_balance):
                    block_purchase.append('block ' + stockticker)
            except Exception as problem:
                print(problem)
                continue

        for element in strong_buy:
            try:
                # for an element in the strong_buy indicator list, if it is equal to the current stock position,
                # then we wont liquidate, if not then we will liquidate
                if element in current_stock_position:
                    continue
                else:
                    # if we have an indicator that is different that we just calculated, we need to remove
                    # the old position and use the new analysis as it is more up to date on the strength of the stock
                    check_orders = api.list_orders(status='open')
                    for order in check_orders:
                        api.cancel_order(order.id)
                    api.close_position(element)
            except Exception as problem:
                print(problem)
                continue

        for element in buy:
            try:
                if element in current_stock_position:
                    continue
                else:
                    check_orders = api.list_orders(status='open')
                    for order in check_orders:
                        api.cancel_order(order.id)
                    api.close_position(element)
            except Exception as problem:
                print(problem)
                continue

        for element in weak_buy:
            try:
                if element in current_stock_position:
                    continue
                else:
                    check_orders = api.list_orders(status='open')
                    for order in check_orders:
                        api.cancel_order(order.id)
                    api.close_position(element)
            except Exception as problem:
                print(problem)
                continue

        for element in strong_sell:
            try:
                if element in current_stock_position:
                    continue
                else:
                    check_orders = api.list_orders(status='open')
                    for order in check_orders:
                        api.cancel_order(order.id)
                    api.close_position(element)
            except Exception as problem:
                print(problem)
                continue

        for element in sell:
            try:
                if element in current_stock_position:
                    continue
                else:
                    check_orders = api.list_orders(status='open')
                    for order in check_orders:
                        api.cancel_order(order.id)
                    api.close_position(element)
            except Exception as problem:
                print(problem)
                continue

        for element in weak_sell:
            try:
                if element in current_stock_position:
                    continue
                else:
                    check_orders = api.list_orders(status='open')
                    for order in check_orders:
                        api.cancel_order(order.id)
                    api.close_position(element)
            except Exception as problem:
                print(problem)
                continue

    #################################################################################################################
    # check orders and see if they should be sold, if we have idle bracket orders with a small fluctuating profit/loss
    # just close them out
    for element in stock_tickers:
        api.list_positions()
        for stockposition in api.list_positions():
            if float(getattr(stockposition, "unrealized_intraday_plpc")) > 0.001:
                check_orders = api.list_orders(status='open')
                for order in check_orders:
                    api.cancel_order(order.id)
                api.close_position(element)

    #################################################################################################################
    # code is necessary to import the c# trading program
    AddReference(r"C:\Users\fabio\source\repos\Main Trade Executor Class Library\Main Trade Executor Class Lib"
                 r"rary\bin\Release\Main Trade Executor Class Library.dll")
    import CSharpTradeExecutor
    trader = CSharpTradeExecutor.BracketOrders()

    print(block_purchase)

    for element in strong_buy:
        try:
            for pos, it in enumerate(block_purchase):
                if element in block_purchase[pos]:
                    raise Exception('{} Position has exceeded 10% of the portfolio value'.format(element))
            price = stock_prices[element]
            account_percentage = (account_balance * 0.04) // price
            round_lot = int(account_percentage)
            if round_lot == 0:
                round_lot += 1
            stop_loss = 0.9985 * price
            stoplosslimitprice = .9980 * price
            limit_price = 1.002 * price
            args = [element, 'buy', str(round_lot), str(round(stop_loss, 2)), str(round(limit_price, 2)),
                    str(round(stoplosslimitprice, 2))]
            trader.Trader(args)
            notification.show_toast("Program Trades Executed", "Program executed strongbuy trade of {} at {}".format(
                                    element, time.strftime("%H:%M:%S")), duration=4)
            pos = str(element + 'strongbuy')
            current_stock_position.append(pos)
        except Exception as error:
            print('There was an {} with the trade execution'.format(error))
            continue

    for element in buy:
        try:
            for pos, it in enumerate(block_purchase):
                if element in block_purchase[pos]:
                    raise Exception('{} Position has exceeded 10% of the portfolio value'.format(element))
            price = stock_prices[element]
            account_percentage = (account_balance * 0.03) // price
            round_lot = int(account_percentage)
            if round_lot == 0:
                round_lot += 1
            limit_price = 1.0016 * price
            stop_loss = 0.9986 * price
            stoplosslimitprice = 0.9984 * price
            args = [element, 'buy', str(round_lot), str(round(stop_loss, 2)), str(round(limit_price, 2)),
                    str(round(stoplosslimitprice, 2))]
            trader.Trader(args)
            notification.show_toast("Program Trades Executed", "Program executed buy trade of {} at {}".format(
                                    element, time.strftime("%H:%M:%S")), duration=4)
            pos = str(element + 'buy')
            current_stock_position.append(pos)
        except Exception as error:
            print('There was an {} with the trade execution'.format(error))
            continue

    for element in weak_buy:
        try:
            for pos, it in enumerate(block_purchase):
                if element in block_purchase[pos]:
                    raise Exception('{} Position has exceeded 10% of the portfolio value'.format(element))
            price = stock_prices[element]
            account_percentage = (account_balance * 0.025) // price
            round_lot = int(account_percentage)
            if round_lot == 0:
                round_lot += 1
            limit_price = 1.0012 * price
            stop_loss = 0.9988 * price
            stoplosslimitprice = 0.9986 * price
            args = [element, 'buy', str(round_lot), str(round(stop_loss, 2)), str(round(limit_price, 2)),
                    str(round(stoplosslimitprice, 2))]
            trader.Trader(args)
            notification.show_toast("Program Trades Executed", "Program executed weakbuy trade of {} at {}".format(
                                    element, time.strftime("%H:%M:%S")), duration=4)
            pos = str(element + 'weakbuy')
            current_stock_position.append(pos)
        except Exception as error:
            print('There was an {} with the trade execution'.format(error))
            continue

    # short trades
    for element in strong_sell:
        try:
            for pos, it in enumerate(block_purchase):
                if element in block_purchase[pos]:
                    raise Exception('{} Position has exceeded 10% of the portfolio value'.format(element))
            price = stock_prices[element]
            account_percentage = (account_balance * 0.04) // price
            round_lot = int(account_percentage)
            if round_lot == 0:
                round_lot += 1
            limit_price = .998 * price
            stop_loss = 1.0015 * price
            stoplosslimitprice = 1.0020 * price
            args = [element, 'sell', str(round_lot), str(round(stop_loss, 2)), str(round(limit_price, 2)),
                    str(round(stoplosslimitprice, 2))]
            trader.Trader(args)
            notification.show_toast("Program Trades Executed", "Program executed strongsell trade of {} at {}".format(
                                    element, time.strftime("%H:%M:%S")), duration=4)
            pos = str(element + 'strongsell')
            current_stock_position.append(pos)
        except Exception as error:
            print('There was an {} with the trade execution'.format(error))
            continue

    for element in sell:
        try:
            for pos, it in enumerate(block_purchase):
                if element in block_purchase[pos]:
                    raise Exception('{} Position has exceeded 10% of the portfolio value'.format(element))
            price = stock_prices[element]
            account_percentage = (account_balance * 0.03) // price
            round_lot = int(account_percentage)
            if round_lot == 0:
                round_lot += 1
            limit_price = .9984 * price
            stop_loss = 1.0012 * price
            stoplosslimitprice = 1.0016 * price
            args = [element, 'sell', str(round_lot), str(round(stop_loss, 2)), str(round(limit_price, 2)),
                    str(round(stoplosslimitprice, 2))]
            trader.Trader(args)
            notification.show_toast("Program Trades Executed", "Program executed sell trade of {} at {}".format(
                                    element, time.strftime("%H:%M:%S")), duration=4)
            pos = str(element + 'sell')
            current_stock_position.append(pos)
        except Exception as error:
            print('There was an {} with the trade execution'.format(error))
            continue

    for element in weak_sell:
        try:
            for pos, it in enumerate(block_purchase):
                if element in block_purchase[pos]:
                    raise Exception('{} Position has exceeded 10% of the portfolio value'.format(element))
            price = stock_prices[element]
            account_percentage = (account_balance * 0.025) // price
            round_lot = int(account_percentage)
            if round_lot == 0:
                round_lot += 1
            limit_price = .9988 * price
            stop_loss = 1.0010 * price
            stoplosslimitprice = 1.0012 * price
            args = [element, 'sell', str(round_lot), str(round(stop_loss, 2)), str(round(limit_price, 2)),
                    str(round(stoplosslimitprice, 2))]
            trader.Trader(args)
            notification.show_toast("Program Trades Executed", "Program executed weaksell trade of {} at {}".format(
                                    element, time.strftime("%H:%M:%S")), duration=4)
            pos = str(element + 'weaksell')
            current_stock_position.append(pos)
        except Exception as error:
            print('There was an {} with the trade execution'.format(error))
            continue


def check_trades():
    if len(trade_data) == 0:
        print('No trades were gathered')
        return False
    return True


def cleanup():
    # get rid of elements in lists that are dated 5 minutes and beyond to save memory
    if (end - start) > 300:
        for element in stock_tickers:
            for p, i in enumerate(trade_data[element]):
                if dt.datetime.strptime(i['time'], "%Y-%m-%d %H:%M:%S.%f") < \
                        dt.datetime.now() - dt.timedelta(seconds=300):
                    trade_data[element].remove(i)
            for po, it in enumerate(quote_data[element]):
                if dt.datetime.strptime(it['time'], "%Y-%m-%d %H:%M:%S") < \
                        dt.datetime.now() - dt.timedelta(seconds=300):
                    quote_data[element].remove(it)
            for pos, ite in enumerate(ti_data[element]):
                if dt.datetime.strptime(ite['time'], "%Y-%m-%d %H:%M:%S") < \
                        dt.datetime.now() - dt.timedelta(seconds=300):
                    ti_data[element].remove(ite)


def check_for_market_close():
    close = dt.datetime.strptime(market_close, '%Y-%m-%d %H:%M:%S.%f')
    if local_timezone == 'EST' and dt.datetime.now() > (close - dt.timedelta(minutes=5)):
        raise Exception('The market is closing in 5 minutes, all positions have been closed')
    if local_timezone == 'CST' and dt.datetime.now() > (close - dt.timedelta(minutes=65)):
        raise Exception('The market is closing in 5 minutes, all positions have been closed')
    if local_timezone == 'CDT' and dt.datetime.now() > (close - dt.timedelta(minutes=65)):
        raise Exception('The market is closing in 5 minutes, all positions have been closed')

"""

END OF ALGO TRADER

"""

def data_to_excel(metrics):
    book = openpyxl.load_workbook('Portfolio Data.xlsx')
    writer = pd.ExcelWriter('Portfolio Data.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    metrics.to_excel(writer, sheet_name=sheet_name)
    try:
        sheet = book['Sheet']
        book.remove(sheet)
        book.save('Portfolio Data.xlsx')
    except KeyError:
        pass
    writer.save()


def formatting_excel(name):
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = 1
    wkb = excel.Workbooks.Open(r"C:\Users\fabio\PycharmProjects\AlgoTrader\Portfolio Data.xlsx")
    ws = wkb.Worksheets(name)
    ws.Columns.AutoFit()
    wkb.Save()
    excel.Application.Quit()


def webscraping(tickers):
    try:
        bond_url = 'http://www.marketwatch.com/investing/bond/tmubmusd01m?countrycode=bx'
        driver.get(bond_url)
        html = driver.execute_script('return document.body.innerHTML;')
        soup = BeautifulSoup(html, 'lxml')
        try:
            bondlist = [entry.text for entry in
                        soup.find_all('bg-quote', {'class': 'value', 'field': 'Last'})]
            risk_free_rate = float(bondlist[0])
            print("1 month risk-free-rate", str(risk_free_rate) + str('%'))
        except IndexError:
            print("Failed to fetch 1-Month T-bond Yield! Setting to default value (0.01)")
            risk_free_rate = 0.01

        for stock in tickers:
            quote = {}
            stock_url = 'https://finance.yahoo.com/quote/' + stock + '?p=' + stock
            driver.get(stock_url)
            html = driver.execute_script('return document.body.innerHTML;')
            soup = BeautifulSoup(html, 'lxml')
            beta_metric = [entry.text for entry in soup.find_all('span', {'data-reactid': '144'})]
            return_pct = [entry.text for entry in soup.find_all('span', {'data-reactid': '51'})]
            delimiter1 = '('
            delimiter2 = ')'
            formatted_return = str(return_pct[0])
            rt = formatted_return[formatted_return.find(delimiter1) + 1: formatted_return.find(delimiter2)]
            return_string = rt.split("%")[0]
            returns = float(return_string)
            print(stock, returns)
            quote['stock'] = stock
            quote['beta'] = beta_metric[0]
            quote['returns'] = round(returns, 4)
            quote_data[stock] = quote
        # spy
        spy_url = 'https://finance.yahoo.com/quote/SPY?p=SPY'
        driver.get(spy_url)
        html = driver.execute_script('return document.body.innerHTML;')
        soup = BeautifulSoup(html, 'lxml')
        spy_returns_pct = [entry.text for entry in soup.find_all('span', {'data-reactid': '51'})]
        delimiter1 = '('
        delimiter2 = ')'
        formatted_return = str(spy_returns_pct[0])
        rt = formatted_return[formatted_return.find(delimiter1) + 1: formatted_return.find(delimiter2)]
        return_string = rt.split("%")[0]
        spy_returns = float(return_string)
        return spy_returns, risk_free_rate
    except Exception as e:
        print(e)
    finally:
        driver.quit()


def purchasing_filter(activities_df):
    # filtering out bull side long purchases
    long_purchases_df = activities_df.loc[(activities_df['side'] == 'buy') & (activities_df['cumulative_sum'] > 0)]
    total_long_purchases = round(long_purchases_df['net_trade'].sum(), 2)
    print("Gross cost of long positions:", total_long_purchases)

    # filtering out bear side 'buy to cover' purchases
    short_purchases_df = activities_df.loc[(activities_df['side'] == 'buy') & (activities_df['cumulative_sum'] <= 0)]
    total_short_purchases = round(short_purchases_df['net_trade'].sum(), 2)
    print("Gross cost of short positions:", total_short_purchases)

    # filtering bull side long sales
    long_sales_df = activities_df.loc[activities_df['side'] == 'sell']
    total_long_sells = round(long_sales_df['net_trade'].sum(), 2)
    print("Gross profit of long positions:", total_long_sells)

    # filtering bear side short purchases
    short_sales_df = activities_df.loc[activities_df['side'] == 'sell_short']
    total_short_sells = round(short_sales_df['net_trade'].sum(), 2)
    print("Gross profit of short positions:", total_short_sells)

    long_buy_df = long_purchases_df.sort_values(['symbol', 'transaction_time'])
    lb_df = pd.DataFrame(long_buy_df)
    lb_df.reset_index(drop=True, inplace=True)

    long_sales_df = long_sales_df.sort_values(['symbol', 'transaction_time'])
    ls_df = pd.DataFrame(long_sales_df)
    ls_df.reset_index(drop=True, inplace=True)

    short_buys_df = short_purchases_df.sort_values(['symbol', 'transaction_time'])
    sb_df = pd.DataFrame(short_buys_df)
    sb_df.reset_index(drop=True, inplace=True)

    short_sells_df = short_sales_df.sort_values(['symbol', 'transaction_time'])
    ss_df = pd.DataFrame(short_sells_df)
    ss_df.reset_index(drop=True, inplace=True)
    return lb_df, ls_df, sb_df, ss_df


def order_settlement():
    current_buy_pos = 0
    current_sell_pos = 0
    ###################################################################################################################
    for place, it in enumerate(short_buy_order_book.copy(), start=current_buy_pos):
        try:
            place = current_buy_pos
            o = 1
            buy_qty = short_buy_order_book[current_buy_pos][2]
            buy_value = short_buy_order_book[current_buy_pos][3]
            if short_buy_order_book[current_buy_pos][1] == 'partial_fill':
                while True:
                    if short_buy_order_book[current_buy_pos + o][1] == 'partial_fill':
                        buy_qty += short_buy_order_book[current_buy_pos + o][2]
                        buy_value += short_buy_order_book[current_buy_pos + o][3]
                        o += 1
                    else:
                        buy_qty += short_buy_order_book[current_buy_pos + o][2]
                        buy_value += short_buy_order_book[current_buy_pos + o][3]
                        short_buy_order_book[current_buy_pos + o][2] = buy_qty
                        short_buy_order_book[current_buy_pos + o][3] = round(buy_value, 2)
                        for j in range(o):
                            del short_buy_order_book[current_buy_pos + j]
                        break
            current_buy_pos += o
            place += o
        except KeyError:
            pass
    print(len(short_buy_order_book), short_buy_order_book)

    for place, it in enumerate(short_sell_order_book.copy(), start=current_sell_pos):
        try:
            place = current_sell_pos
            o = 1
            buy_qty = short_sell_order_book[current_sell_pos][2]
            buy_value = short_sell_order_book[current_sell_pos][3]
            if short_sell_order_book[current_sell_pos][1] == 'partial_fill':
                while True:
                    if short_sell_order_book[current_sell_pos + o][1] == 'partial_fill':
                        buy_qty += short_sell_order_book[current_sell_pos + o][2]
                        buy_value += short_sell_order_book[current_sell_pos + o][3]
                        o += 1
                    else:
                        buy_qty += short_sell_order_book[current_sell_pos + o][2]
                        buy_value += short_sell_order_book[current_sell_pos + o][3]
                        short_sell_order_book[current_sell_pos + o][2] = buy_qty
                        short_sell_order_book[current_sell_pos + o][3] = round(buy_value, 2)
                        for j in range(o):
                            del short_sell_order_book[current_sell_pos + j]
                        break
            current_sell_pos += o
            place += o
        except KeyError:
            pass
    print(len(short_sell_order_book), short_sell_order_book)

    current_buy_pos = 0
    current_sell_pos = 0
    for place, it in enumerate(buy_order_book.copy(), start=current_buy_pos):
        try:
            place = current_buy_pos
            o = 1
            buy_qty = buy_order_book[current_buy_pos][2]
            buy_value = buy_order_book[current_buy_pos][3]
            if buy_order_book[current_buy_pos][1] == 'partial_fill':
                while True:
                    if buy_order_book[current_buy_pos + o][1] == 'partial_fill':
                        buy_qty += buy_order_book[current_buy_pos + o][2]
                        buy_value += buy_order_book[current_buy_pos + o][3]
                        o += 1
                    else:
                        buy_qty += buy_order_book[current_buy_pos + o][2]
                        buy_value += buy_order_book[current_buy_pos + o][3]
                        buy_order_book[current_buy_pos + o][2] = buy_qty
                        buy_order_book[current_buy_pos + o][3] = round(buy_value, 2)
                        for h in range(o):
                            del buy_order_book[current_buy_pos + h]
                        break
            current_buy_pos += o
            place += o
        except KeyError:
            pass
    print(len(buy_order_book), buy_order_book)

    for place, it in enumerate(sell_order_book.copy(), start=current_sell_pos):
        try:
            place = current_sell_pos
            o = 1
            sell_qty = sell_order_book[current_sell_pos][2]
            sell_value = sell_order_book[current_sell_pos][3]
            if sell_order_book[current_sell_pos][1] == 'partial_fill':
                while True:
                    if sell_order_book[current_sell_pos + o][1] == 'partial_fill':
                        sell_qty += sell_order_book[current_sell_pos + o][2]
                        sell_value += sell_order_book[current_sell_pos + o][3]
                        o += 1
                    else:
                        sell_qty += sell_order_book[current_sell_pos + o][2]
                        sell_value += sell_order_book[current_sell_pos + o][3]
                        sell_order_book[current_sell_pos + o][2] = sell_qty
                        sell_order_book[current_sell_pos + o][3] = round(sell_value, 2)
                        for p in range(o):
                            del sell_order_book[current_sell_pos + p]
                        break
            current_sell_pos += o
            place += o
        except KeyError:
            pass
    print(len(sell_order_book), sell_order_book)


def trade_book_settlement():
    trade_book = {}
    short_trade_book = {}
    short_order_time_held = {}
    long_order_time_held = {}
    ####################################################################################################################
    # adding to respective trade books
    current_buy_pos = 0
    current_sell_pos = 0
    short_trade_ledger_position = 0

    while len(short_buy_order_book) > 0:
        for pos, k in enumerate(short_buy_order_book.copy(), start=current_buy_pos):
            d1 = dt.datetime.strptime(short_buy_order_book[current_buy_pos][0], "%Y-%m-%dT%H:%M:%S.%fZ")
            d2 = dt.datetime.strptime(short_sell_order_book[current_sell_pos][0], "%Y-%m-%dT%H:%M:%S.%fZ")
            elapsed_time = (d1 - d2).total_seconds()

            if short_buy_order_book[current_buy_pos][5] == short_sell_order_book[current_sell_pos][5]:
                if short_buy_order_book[current_buy_pos][2] == short_sell_order_book[current_sell_pos][2]:
                    short_trade_book[short_trade_ledger_position] = round(short_buy_order_book[current_buy_pos][3] +
                                                                          short_sell_order_book[current_sell_pos][3], 2)
                    short_order_time_held[short_trade_ledger_position] = (elapsed_time,
                                                                          short_buy_order_book[current_buy_pos][2])
                    del short_buy_order_book[current_buy_pos], short_sell_order_book[current_sell_pos]

                else:
                    while True:
                        bought_share_value = round(
                            short_buy_order_book[current_buy_pos][3] / short_buy_order_book[current_buy_pos][2], 2)
                        sold_share_value = round(
                            short_sell_order_book[current_sell_pos][3] / short_sell_order_book[current_sell_pos][2], 2)
                        buy_quantity = short_buy_order_book[current_buy_pos][2]
                        sell_quantity = short_sell_order_book[current_sell_pos][2]
                        buy_val = short_buy_order_book[current_buy_pos][3]
                        sell_val = short_sell_order_book[current_sell_pos][3]

                        if buy_quantity > sell_quantity:
                            value_of_shares_sold = bought_share_value * sell_quantity
                            short_trade_book[short_trade_ledger_position] = round(value_of_shares_sold + sell_val, 2)
                            short_buy_order_book[current_buy_pos][2] -= sell_quantity
                            short_buy_order_book[current_buy_pos][3] = round(buy_val - value_of_shares_sold, 2)
                            short_buy_order_book[current_buy_pos][4] -= sell_quantity
                            short_order_time_held[short_trade_ledger_position] = (elapsed_time, sell_quantity)
                            del short_sell_order_book[current_sell_pos]
                            break

                        if buy_quantity < sell_quantity:
                            value_of_shares_sold = sold_share_value * buy_quantity
                            short_sell_order_book[current_sell_pos][2] -= buy_quantity
                            short_sell_order_book[current_sell_pos][3] = round(sell_val - value_of_shares_sold, 2)
                            short_trade_book[short_trade_ledger_position] = round(value_of_shares_sold + buy_val, 2)
                            short_order_time_held[short_trade_ledger_position] = (elapsed_time, buy_quantity)
                            del short_buy_order_book[current_buy_pos]
                        break
            else:
                price_per_share = round(short_sell_order_book[current_sell_pos][3] /
                                        short_sell_order_book[current_sell_pos][2], 2)
                print("There is a net short position of {} shares on {} at a price of ${} per share".format(
                    short_sell_order_book[current_sell_pos][2], short_sell_order_book[current_sell_pos][5],
                    price_per_share))
                del short_sell_order_book[current_sell_pos]

            if len(short_buy_order_book) > 0:
                current_buy_pos = list(short_buy_order_book)[0]
                current_sell_pos = list(short_sell_order_book)[0]
                short_trade_ledger_position += 1

    current_buy_pos = 0
    current_sell_pos = 0
    trade_ledger_position = 0

    while len(buy_order_book) > 0:
        for pos, k in enumerate(buy_order_book.copy(), start=current_buy_pos):
            d1 = dt.datetime.strptime(buy_order_book[current_buy_pos][0], "%Y-%m-%dT%H:%M:%S.%fZ")
            d2 = dt.datetime.strptime(sell_order_book[current_sell_pos][0], "%Y-%m-%dT%H:%M:%S.%fZ")
            elapsed_time = (d2 - d1).total_seconds()
            # check if the symbols match

            if buy_order_book[current_buy_pos][5] == sell_order_book[current_sell_pos][5]:
                if buy_order_book[current_buy_pos][2] == sell_order_book[current_sell_pos][2]:
                    trade_book[trade_ledger_position] = round(
                        buy_order_book[current_buy_pos][3] + sell_order_book[current_sell_pos][3], 2)
                    long_order_time_held[trade_ledger_position] = (elapsed_time, buy_order_book[current_buy_pos][2])
                    del buy_order_book[current_buy_pos], sell_order_book[current_sell_pos]

                else:
                    while True:
                        bought_share_value = round(buy_order_book[current_buy_pos][3] / buy_order_book[current_buy_pos][2],
                                                   2)
                        sold_share_value = round(
                            sell_order_book[current_sell_pos][3] / sell_order_book[current_sell_pos][2], 2)
                        buy_quantity = buy_order_book[current_buy_pos][2]
                        sell_quantity = sell_order_book[current_sell_pos][2]
                        buy_val = buy_order_book[current_buy_pos][3]
                        sell_val = sell_order_book[current_sell_pos][3]

                        if buy_quantity > sell_quantity:
                            value_of_shares_sold = bought_share_value * sell_quantity
                            trade_book[trade_ledger_position] = round(value_of_shares_sold + sell_val, 2)
                            buy_order_book[current_buy_pos][2] -= sell_quantity
                            buy_order_book[current_buy_pos][3] = round(buy_val - value_of_shares_sold, 2)
                            buy_order_book[current_buy_pos][4] -= sell_quantity
                            long_order_time_held[trade_ledger_position] = (elapsed_time, sell_quantity)
                            del sell_order_book[current_sell_pos]
                            break

                        if buy_quantity < sell_quantity:
                            value_of_shares_sold = sold_share_value * buy_quantity
                            sell_order_book[current_sell_pos][2] -= buy_quantity
                            sell_order_book[current_sell_pos][3] = round(sell_val - value_of_shares_sold, 2)
                            trade_book[trade_ledger_position] = round(value_of_shares_sold + buy_val, 2)
                            long_order_time_held[trade_ledger_position] = (elapsed_time, buy_quantity)
                            del buy_order_book[current_buy_pos]
                        break
            else:
                price_per_share = round(buy_order_book[current_buy_pos][3] / buy_order_book[current_buy_pos][2], 2)
                print("There is a net long position of {} shares on {} at a price of ${} per share".format(
                    buy_order_book[current_buy_pos][2], buy_order_book[current_buy_pos][5], price_per_share))
                del buy_order_book[current_buy_pos]

            if len(buy_order_book) > 0:
                current_buy_pos = list(buy_order_book)[0]
                current_sell_pos = list(sell_order_book)[0]
                trade_ledger_position += 1
    return short_trade_book, trade_book, short_order_time_held, long_order_time_held


if __name__ == '__main__':
    if not os.path.isfile("Ticker Selections.xlsx"):
        wb = openpyxl.Workbook()
        wb.save('Ticker Selections.xlsx')

    try:
        shutil.rmtree("Daily Stock Analysis/Stocks")
        os.mkdir("Daily Stock Analysis/Stocks")
    except Exception as e:
        print(e)
        pass
    ###################################################################################################################
    key = "PKCPC6RJ84BG84W3PB60"
    sec = "U1r9Z2QknL9FwAaTztfLl5g1DTxpa5m97qyWCGZ7"
    url = "https://paper-api.alpaca.markets"
    api = trade_api.REST(key, sec, url, api_version='v2')
    stock_tickers = get_tickers()
    api_calls(stock_tickers)
    obv_score_array()
    notification = ToastNotifier()
    websocket.enableTrace(True)
    websocket_url = "wss://ws.finnhub.io?token=bsq43v0fkcbdt6un6ivg"
    socket = websocket.WebSocketApp(websocket_url, on_message=on_message, on_error=on_error, on_close=on_close)
    socket.on_open = on_open
    account = api.get_account()
    clock = api.get_clock()
    ###################################################################################################################
    account_balance = float(account.buying_power) / 4
    print('Trading Account status:', account.status)
    date = dt.datetime.date(dt.datetime.now(dt.timezone.utc))
    # Get when the market opens or opened today
    fulltimezone = str(dt.datetime.now(dt.timezone.utc).astimezone().tzinfo)
    local_timezone = ''.join([c for c in fulltimezone if c.isupper()])
    print('The date is:', str(dt.date.today()), 'The time is', str(dt.datetime.now().time()), local_timezone)
    market_close = str(clock.next_close)[:16:] + str(':00.000000')
    stock_tickers = []
    while True:
        minimum_position = 1.0
        maximum_position = 3.0
        retry = False
        start_reducing = False
        stock_tickers = []
        try:
            if os.path.isfile('Daily Stock Analysis/OBV_Ranked.csv'):
                if os.stat('Daily Stock Analysis/OBV_Ranked.csv').st_size > 0:
                    with open('Daily Stock Analysis/OBV_Ranked.csv', 'r') as f:
                        reader = csv.reader(f)
                        for position, row in enumerate(reader):
                            if position > 0:
                                if minimum_position <= position <= maximum_position:
                                    if row[0] not in stock_tickers:
                                        stock_tickers.append(row[0])
            print(stock_tickers)
            AddReference(r"C:\Users\fabio\source\repos\Webscraper Class Library\Webscraper Class Library\bin\Debug\Web"
                         r"scraper Class Library.dll")
            import CSharpwebscraper

            scraper = CSharpwebscraper.Webscraper()
            for position, item in enumerate(stock_tickers):
                stock = [stock_tickers[position]]
                try:
                    stock_info = scraper.Scraper(stock)
                except Exception as e:
                    print(e, "Error Found with Tickers!")
                    line = stock_tickers[position]
                    lines = []
                    with open('Daily Stock Analysis/OBV_Ranked.csv', 'r') as f:
                        reader = csv.reader(f)
                        for place, row in enumerate(reader):
                            if place > 0:
                                if row[0] == line:
                                    start_reducing = True
                                    continue
                                if start_reducing:
                                    row[2] = float(row[2]) - 1
                                lines.append(row)
                            else:
                                lines.append(row)
                    with open('Daily Stock Analysis/OBV_Ranked.csv', 'w') as w:
                        writer = csv.writer(w, lineterminator='\n')
                        writer.writerows(lines)
                    retry = True
            if not retry:
                break
        except Exception as e:
            print(e)
            print("An error with the automated stock fetcher was found, using the manual stock fetcher")
            stock_tickers = []
            break

    print(stock_tickers)
    if len(stock_tickers) == 0:
        # manual
        print("The market closes at {} o'clock EST today.".format(market_close[11:16:]))
        print("Input stock tickers separated by a space, the quotes and trades for each stock will be streamed")
        print("When you are done entering tickers, press Enter to show the quotes for each stock in order")
        print("Enter in 'close' in order to close all current positions")
        stock_tickers = input('Enter Ticker(s): ').upper().split()

        while True:
            try:
                if stock_tickers == ['CLOSE']:
                    api.cancel_all_orders()
                    api.close_all_positions()
                    stock_tickers = input('Positions have been closed, Enter Ticker(s): ').upper().split()

                for position, item in enumerate(stock_tickers):
                    try:
                        asset = api.get_asset(item)
                        if not asset.tradable:
                            print(item, 'is not available to trade on Alpaca!')
                            stock_tickers[position] = input('Enter different ticker: ').upper()
                        continue
                    except Exception as e:
                        print(e)
                        print(stock_tickers[position], 'is not a valid ticker!')
                        stock_tickers[position] = input('Enter a different ticker: ').upper()

                for stock in stock_tickers:
                    try:
                        asset = api.get_asset(stock)
                        if not asset.tradable:
                            raise Exception("Not Tradable")
                    except Exception:
                        raise Exception("Not Tradable")
                break
            except Exception as stockinputerror:
                print(stockinputerror)
                print("There was a problem with the ticker(s) that you entered")
                continue
    else:
        pass

    while True:
        start = time.time()
        stock_buylist = {}
        stock_shortlist = {}
        stock_prices = {}
        volume_terms = {}
        trade_data = {}
        ti_data = {}
        quote_data = {}
        stock_list_length = {}
        indicator_votes = {}
        current_stock_position = []
        stock_price_movement = {}
        cutoff_bool = False
        for ticker in stock_tickers:
            indicator_votes[ticker] = {'Bullish Votes': 0, 'Bearish Votes': 0, 'Neutral Votes': 0}
            trade_data[ticker] = []
            ti_data[ticker] = []
            quote_data[ticker] = []
            stock_buylist[ticker] = []
            stock_shortlist[ticker] = []
            stock_price_movement[ticker] = ''
        ############################################################################################################
        while True:
            try:
                strong_buy = []
                buy = []
                weak_buy = []
                weak_sell = []
                sell = []
                strong_sell = []
                tradethread = th.Thread(target=trade_execution_operations)
                tradethread.daemon = True
                check_for_market_close()
                main_data_engine()
                print('Trades:', trade_data)
                print('Quotes:', quote_data)
                print('Indicators:', ti_data)
                if not check_trades():
                    print('Warning, No trades gathered! Program terminating...')
                    raise SystemExit(0)
                ##############################
                ticker_operations()
                volume_operations()
                analysis_operations()
                tradethread.start()
                end = time.time()
                print('Time Elapsed (in seconds):', int((end - start)))
                cleanup()
                tradethread.join()
            except Exception('The market is closing in 5 minutes, all positions have been closed') as e:
                print(e)
                cutoff_bool = True
                api.close_all_positions()
                api.cancel_all_orders()
                break
            """
            except Exception as e:
                notification.show_toast("Program Critical Error", "Program Raised Error {}".format(e), duration=5)
                print('All pending orders will be cancelled and all positions will be liquidated immediately')
                api.cancel_all_orders()
                api.close_all_positions()
            """
        if cutoff_bool:
            break


    # START OF PORTFOLIO ANALYSIS
    if not os.path.isfile(r"C:\Users\fabio\PycharmProjects\AlgoTrader\Portfolio Data.xlsx"):
        wb = openpyxl.Workbook()
        wb.save('Portfolio Data.xlsx')

    pd.options.mode.chained_assignment = None

    # Can also limit the results by date if desired.
    days = 1
    while True:
        try:
            spec_date = dt.datetime.today() - dt.timedelta(days=days)
            date = spec_date.strftime('%Y-%m-%d')
            activities = api.get_activities(activity_types='FILL', date=date)
            activities_df = pd.DataFrame([activity._raw for activity in activities])
            if not len(activities_df) > 50:
                raise Exception("Not enough trades to not be a test")
            else:
                print("Analyzing Portfolio Activities on {}".format(date))

            activities_df = activities_df.iloc[::-1]
            activities_df[['price', 'qty']] = activities_df[['price', 'qty']].apply(pd.to_numeric)
            activities_df['net_qty'] = np.where(activities_df.side == 'buy', activities_df.qty, -activities_df.qty)
            activities_df['net_trade'] = -activities_df.net_qty * activities_df.price
            activities_df['cumulative_sum'] = activities_df.groupby('symbol')['net_qty'].apply(lambda g: g.cumsum())
            break
        except Exception as e:
            days += 1

    activities_df.to_excel("Portfolio Activities.xlsx")

    ###################################################################################################################
    # Total Net Profit for Long and Short Trades
    long_buy_df, long_sell_df, short_buy_df, short_sell_df = purchasing_filter(activities_df)

    # we can make an order book that tracks each trade as it iterates down the list
    short_buy_order_book = {}
    for index, row in short_buy_df.iterrows():
        short_buy_order_book[index] = [row['transaction_time'], row['type'], row['qty'], round(row['net_trade'], 2),
                                       row['cumulative_sum'], row['symbol']]
    print("short buys", len(short_buy_order_book), short_buy_order_book)
    short_sell_order_book = {}
    for index, row in short_sell_df.iterrows():
        short_sell_order_book[index] = [row['transaction_time'], row['type'], row['qty'], round(row['net_trade'], 2),
                                        row['cumulative_sum'], row['symbol']]
    print("short sells", len(short_sell_order_book), short_sell_order_book)
    buy_order_book = {}
    for index, row in long_buy_df.iterrows():
        buy_order_book[index] = [row['transaction_time'], row['type'], row['qty'], round(row['net_trade'], 2),
                                 row['cumulative_sum'], row['symbol']]
    print("long buys", len(buy_order_book), buy_order_book)
    sell_order_book = {}
    for index, row in long_sell_df.iterrows():
        sell_order_book[index] = [row['transaction_time'], row['type'], row['qty'], round(row['net_trade'], 2),
                                  row['cumulative_sum'], row['symbol']]
    print("long sells", len(sell_order_book), sell_order_book)

    order_settlement()
    short_trades, long_trades, short_trade_time_ledger, long_trade_time_ledger = trade_book_settlement()
    ###################################################################################################################
    # finding average short trade held time
    # print(short_trade_time_ledger)
    # print(long_trade_time_ledger)
    short_hold_time = 0
    for position, item in enumerate(short_trade_time_ledger):
        try:
            short_hold_time += short_trade_time_ledger[position][0]
        except KeyError:
            pass
    average_short_trade_hold_time = round(short_hold_time / len(short_trade_time_ledger), 2)
    avg_total_trade_hold_time = average_short_trade_hold_time
    avg_short_trade_length = time.strftime("%M:%S", time.gmtime(average_short_trade_hold_time))
    # print(avg_short_trade_length)

    long_hold_time = 0
    for position, item in enumerate(long_trade_time_ledger):
        long_hold_time += long_trade_time_ledger[position][0]
    average_long_trade_hold_time = round(long_hold_time / len(long_trade_time_ledger), 2)
    avg_total_trade_hold_time += average_long_trade_hold_time
    avg_long_trade_length = time.strftime("%M:%S", time.gmtime(average_long_trade_hold_time))
    # print(avg_long_trade_length)
    # divide time held in seconds by time of avg trading day, 23400 seconds
    avg_total_trade_hold_time = time.strftime("%M:%S", time.gmtime((avg_total_trade_hold_time / 2)))
    ##################################################################
    total_gross_profit = 0
    total_gross_loss = 0
    short_gross_profit = 0
    short_gross_loss = 0
    net_short_profit = 0
    total_short_trades = 0
    short_winning_trades = 0
    short_even_trades = 0
    short_losing_trades = 0
    for i in range(len(short_trades)):
        try:
            if short_trades[i] > 0:
                short_winning_trades += 1
                short_gross_profit += short_trades[i]
                total_gross_profit += short_trades[i]
            elif short_trades[i] < 0:
                short_losing_trades += 1
                short_gross_loss += short_trades[i]
                total_gross_loss += short_trades[i]
            else:
                short_even_trades += 1
            total_short_trades += 1
            net_short_profit += short_trades[i]
        except KeyError:
            pass
    net_short_profit = round(net_short_profit, 2)
    print("Short-side net profit:", net_short_profit)
    print("Short-side profitable trades:", short_winning_trades)
    print("Short-side even trades:", short_even_trades)
    print("Short-side Losing trades:", short_losing_trades)
    print("Total short-side trades:", total_short_trades)

    # initialization of long variables
    long_gross_profit = 0
    long_gross_loss = 0
    net_long_profit = 0
    total_long_trades = 0
    long_winning_trades = 0
    long_even_trades = 0
    long_losing_trades = 0
    for i in range(len(long_trades)):
        if long_trades[i] > 0:
            long_winning_trades += 1
            long_gross_profit += long_trades[i]
            total_gross_profit += long_trades[i]
        elif long_trades[i] < 0:
            long_losing_trades += 1
            long_gross_loss += long_trades[i]
            total_gross_loss += long_trades[i]
        else:
            long_even_trades += 1
        total_long_trades += 1
        net_long_profit += long_trades[i]
    net_long_profit = round(net_long_profit, 2)
    print("\nLong-side net profit:", net_long_profit)
    print("Long-side profitable trades:", long_winning_trades)
    print("Long-side even trades:", long_even_trades)
    print("Long-side losing trades:", long_losing_trades)
    print("Total long-side trades", total_long_trades)

    avg_winning_trade = round((total_gross_profit / (long_winning_trades + short_winning_trades)), 2)
    avg_losing_trade = round((total_gross_loss / (total_long_trades + short_losing_trades)), 2)
    avg_long_winning_trade = round(long_gross_profit / long_winning_trades, 2)
    avg_long_losing_trade = round(long_gross_loss / long_losing_trades, 2)
    avg_short_winning_trade = round(short_gross_profit / short_winning_trades, 2)
    avg_short_losing_trade = round(short_gross_loss / short_losing_trades, 2)

    todayspandl = round(total_gross_profit + total_gross_loss, 2)
    total_gross_profit = round(total_gross_profit, 2)
    total_gross_loss = round(total_gross_loss, 2)
    print("\nProfit Metrics:")
    print("Gross Profit:", total_gross_profit)
    print("Average Winning Trade:", avg_winning_trade)
    print("Gross Loss:", total_gross_loss)
    print("Average Losing Trade:", avg_losing_trade)
    print("Total Net Profit:", todayspandl)

    # profit per symbol
    # handle having outstanding positions and remove the net positive positions so we can pin down the net gain
    # and loss for the stock on the trading day
    non_zero_trades = activities_df.groupby('symbol').filter(lambda trade: sum(trade.net_qty) != 0)
    # print(non_zero_trades)
    # the last row of the dataframe will have potentially the outstanding position
    # print(non_zero_trades.iloc[-1])
    while non_zero_trades.iloc[-1][13] != 0:
        print("Details on outstanding position(s):")
        print(non_zero_trades.iloc[-1])
        dfindex = non_zero_trades.index[-1]
        non_zero_trades = non_zero_trades.drop(dfindex)
        activities_df = activities_df.drop(dfindex)

    net_zero_trades = activities_df.groupby('symbol').filter(lambda trade: sum(trade.net_qty) == 0)
    trades = net_zero_trades.groupby('symbol').net_trade
    profit_per_symbol = net_zero_trades.groupby('symbol').net_trade.sum()
    print("Net Profit per stock:")
    print(profit_per_symbol)

    stock_tickers_involved = profit_per_symbol.index.tolist()
    # print(stock_tickers_involved)
    #############################################################################
    pd.options.display.float_format = '{:.0f}'.format
    quote_data = {}
    for stock in stock_tickers_involved:
        quote_data[stock] = []
    #########################################################################################
    # for quick debugging
    driver = webdriver.Chrome(ChromeDriverManager().install())
    spyreturn, riskfreerate = webscraping(stock_tickers_involved)
    spyreturn = '{:.4f}'.format(spyreturn)
    spyreturn = float(spyreturn)
    print(quote_data)
    print(spyreturn)
    #############################################################################
    account = api.get_account()
    stock_metrics = [['' for m in range(1)] for i in range(len(stock_tickers_involved) * 3)]
    spdr_string = str(spyreturn) + str('%')
    spdr_list = ["Daily return of $SPY", spdr_string]

    stock_index = 0
    for stock in stock_tickers_involved:
        # the average max holdings of each stock is limited at 10% of the portfolio
        trade_size_relative_to_portfolio = 0.1
        beta = trade_size_relative_to_portfolio * float(quote_data[stock]['beta'])
        buying_power = float(account.buying_power) / 4
        stock_profit_pct = round((profit_per_symbol[stock] / buying_power) * 100, 4)
        market_returns_pct = quote_data[stock]['returns']
        # divide the 1 month risk free rate by 30 to approximate the rate of bond return for 1 day
        alpha = round((stock_profit_pct - (riskfreerate / 30)) - (beta * (spyreturn - (riskfreerate / 30))), 4)

        list1 = ["Performance of {}:".format(stock), str(market_returns_pct) + str('%')]
        list2 = ["Performance of {} relative to $SPY:".format(stock), str(round(market_returns_pct - spyreturn, 2)) + str('%')]
        list3 = ["\"Alpha\" trading performance of {}:".format(stock), str(alpha) + str('%')]

        stock_metrics[stock_index] = list1
        stock_metrics[stock_index + 1] = list2
        stock_metrics[stock_index + 2] = list3

        print(list1)
        print(list2)
        print(list3)
        stock_index += 3

    # if the total gross profit of the stock is not positive then the gross loss must be flipped in order to generate a
    # non negative percentage for the profit factor of the portfolio
    if total_gross_profit > -total_gross_loss:
        total_profit_factor = round((total_gross_profit / -total_gross_loss), 2)
    else:
        total_profit_factor = round((total_gross_profit / total_gross_loss), 2)

    if long_gross_profit > -long_gross_loss:
        long_profit_factor = round((long_gross_profit / -long_gross_loss), 2)
    else:
        long_profit_factor = round((long_gross_profit / long_gross_loss), 2)

    if short_gross_profit > -short_gross_loss:
        short_profit_factor = round((short_gross_profit / -short_gross_loss), 2)
    else:
        short_profit_factor = round((short_gross_profit / short_gross_loss), 2)

    total_percent_profitable = round(((long_winning_trades + short_winning_trades) / (total_long_trades +
                                                                                      total_short_trades)) * 100, 2)
    long_percent_profitable = round((long_winning_trades / total_long_trades) * 100, 2)
    short_percent_profitable = round((short_winning_trades / total_short_trades) * 100, 2)

    no_of_stock_metrics = len(stock_metrics)

    # this is the 2d list used to convert into a pandas dataframe for easy transcription onto an excel document
    data = [['Profit Metrics', '', '', ''],
        ['Total Net Profit:', todayspandl, net_long_profit, net_short_profit],
        ['Gross Profit:', total_gross_profit, long_gross_profit, short_gross_profit],
        ['Gross Loss:', total_gross_loss, long_gross_loss, short_gross_loss],
        ['Profit Factor:', total_profit_factor, long_profit_factor, short_profit_factor],
        ['', '', '', ''],
        ['Trade Metrics', '', '', ''],
        ['Total Number of Trades:', int(total_long_trades + total_short_trades), total_long_trades, total_short_trades],
        ['Percent Profitable:', str(total_percent_profitable) + str("%"), str(long_percent_profitable) + str('%'),
         str(short_percent_profitable) + str('%')],
        ['Average Time Held (Minutes):', avg_total_trade_hold_time.lstrip("0"), avg_long_trade_length.lstrip("0"),
         avg_short_trade_length.lstrip("0")],
        ['Winning Trades:', long_winning_trades + short_winning_trades, long_winning_trades, short_winning_trades],
        ['Average Winning Trade:', avg_winning_trade, avg_long_winning_trade, avg_short_winning_trade],
        ['Losing Trades:', long_losing_trades + short_losing_trades, long_losing_trades, short_losing_trades],
        ['Average Losing Trade:', avg_losing_trade, avg_long_losing_trade, avg_short_losing_trade],
        ['Even Trades', long_even_trades + short_even_trades, long_even_trades, short_even_trades],
        ['', '', '', ''],
        ['Stock Metrics', '', '', ''],
        spdr_list] + stock_metrics

    # if i want to add further rows in the future, simply do
    # data + 2d list

    portfolio_metrics = DataFrame(data, columns=['Performance Summary', 'All Trades', 'Long Trades', 'Short Trades'])
    sheet_name = str('Performance on ') + str(date)
    data_to_excel(portfolio_metrics)
    formatting_excel(sheet_name)
