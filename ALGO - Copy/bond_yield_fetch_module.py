from requests import get
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
import os
import glob
import datetime as dt
from dateutil.parser import parse

from ALGO.excel_formatting_module import ExcelFormatting

"""
Replace the government bond yield curve (which is quite wrong to use) and either use the LIBOR
rates or the OIS (Overnight Index Swap) rates
https://www-2.rotman.utoronto.ca/~hull/DownloadablePublications/LIBORvsOIS.pdf
Traditionally practitioners have used LIBOR and LIBOR-swap rates as proxies for risk-free rates
when valuing derivatives. This practice has been called into question by the credit crisis that
started in 2007. Many banks now consider that overnight indexed swap (OIS) rates should be
used as the risk-free rate when collateralized portfolios are valued and that LIBOR should be
used for this purpose when portfolios are not collateralized. This paper examines this practice
and concludes that OIS rates should be used in all situations.
"""


class bondYields:
    def __init__(self):
        self.bond_yields = {}
        self.todays_date = dt.datetime.today()
        self.r = get("https://www.treasury.gov/resource-center/data-chart-center/interest-rates/Pages/Text"
                     "View.aspx?data=yield")
        self.r1 = get("https://www.global-rates.com/en/interest-rates/libor/american-dollar/american-dollar.aspx")
        self.cwd = os.getcwd()
        bond_files = glob.glob(self.cwd + r"\Daily Stock Analysis\Bonds\*.xlsx")
        for i in bond_files:
            os.remove(i)

    # LIBOR rates are okay for pricing derivatives
    def LIBOR_yields(self):
        soup = BeautifulSoup(self.r1.text, 'lxml')
        libor_list = [entry.text for entry in soup.find_all('td', {'align': 'center'})]

        libor_2dlist = []
        row = []
        date = ''
        for i in range(1, 81):
            if i == 1:
                date = libor_list[i]
            row.append(libor_list[i])
            if i % 5 == 0:
                libor_2dlist.append(row)
                row = []

        libor_df = pd.DataFrame.from_records(libor_2dlist)
        libor_refitted_df = libor_df.set_axis(libor_2dlist[0], axis='columns')
        libor_refitted_df = libor_refitted_df.iloc[1:, :]
        terms = ['USD-LIBOR-OVERNIGHT', 'USD-LIBOR-1-WEEK', 'USD-LIBOR-2-WEEKS', 'USD-LIBOR-1-MONTH',
                 'USD-LIBOR-2-MONTHS', 'USD-LIBOR-3-MONTHS', 'USD-LIBOR-4-MONTHS', 'USD-LIBOR-5-MONTHS',
                 'USD-LIBOR-6-MONTHS', 'USD-LIBOR-7-MONTHS', 'USD-LIBOR-8-MONTHS', 'USD-LIBOR-9-MONTHS',
                 'USD-LIBOR-10-MONTHS', 'USD-LIBOR-11-MONTHS', 'USD-LIBOR-12-MONTHS']
        libor_refitted_df.index = terms

        url = self.cwd + fr"\Daily Stock Analysis\Bonds\LIBOR Yields (last updated {date}).xlsx"
        wb = openpyxl.Workbook()
        wb.save(url)
        book = openpyxl.load_workbook(url)
        writer = pd.ExcelWriter(url, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        libor_refitted_df.to_excel(writer, sheet_name=f"LIBOR yields for {date}")
        try:
            sheet = book['Sheet']
            book.remove(sheet)
        except KeyError:
            pass
        writer.save()
        writer.close()
        book.save(url)
        book.close()

        ExcelFormatting(file_path=url).formatting()

        overnight_rate = float(libor_refitted_df.iloc[0, 0].split()[0])
        one_week_rate = float(libor_refitted_df.iloc[1, 0].split()[0])
        one_month_rate = float(libor_refitted_df.iloc[3, 0].split()[0])
        two_month_rate = float(libor_refitted_df.iloc[4, 0].split()[0])
        three_month_rate = float(libor_refitted_df.iloc[5, 0].split()[0])
        six_month_rate = float(libor_refitted_df.iloc[8, 0].split()[0])
        twelve_month_rate = float(libor_refitted_df.iloc[14, 0].split()[0])

        libor_yields = (round(overnight_rate, 6), round(one_week_rate, 6), round(one_month_rate, 6),
                        round(two_month_rate, 6), round(three_month_rate, 6), round(six_month_rate, 6),
                        round(twelve_month_rate, 6))
        return libor_yields

    def treasury_bond_yields(self):
        soup = BeautifulSoup(self.r.text, 'lxml')
        bond_list = [entry.text for entry in soup.find_all('td', {'class': 'text_view_data'})]
        date = None
        for i in range(len(bond_list)):
            if i % 13 == 0:
                date = bond_list[i]
                self.bond_yields[date] = []
                continue
            self.bond_yields[date].append(bond_list[i])

        bond_df = pd.DataFrame.from_dict(self.bond_yields).T
        bond_df = bond_df.rename(columns={0: '1 Mo', 1: '2 Mo', 2: '3 Mo', 3: '6 Mo', 4: '1 Yr', 5: '2 Yr',
                                          6: '3 Yr', 7: '5 Yr', 8: '7 Yr', 9: '10 Yr', 10: '20 Yr', 11: '30 Yr'})
        date_string = date.replace("/", '-')

        parsed_date = parse(date_string).date()

        url = self.cwd + f"\\Daily Stock Analysis\\Bonds\\US T-Bond Yields (last updated {parsed_date}).xlsx"
        wb = openpyxl.Workbook()
        wb.save(url)
        book = openpyxl.load_workbook(url)
        writer = pd.ExcelWriter(url, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        bond_df.to_excel(writer, sheet_name=f"Treasury Bonds for {parsed_date}")
        try:
            sheet = book['Sheet']
            book.remove(sheet)
        except KeyError:
            pass
        writer.save()
        writer.close()
        book.save(url)
        book.close()

        ExcelFormatting(file_path=url).formatting()

        onemonthyield = float(bond_df.iloc[-1, 0]) / 100
        twomonthyield = float(bond_df.iloc[-1, 1]) / 100
        threemonthyield = float(bond_df.iloc[-1, 2]) / 100
        sixmonthyield = float(bond_df.iloc[-1, 3]) / 100
        oneyryield = float(bond_df.iloc[-1, 4]) / 100
        twoyryield = float(bond_df.iloc[-1, 5]) / 100
        return (round(onemonthyield, 6), round(twomonthyield, 6), round(threemonthyield, 6),
                round(sixmonthyield, 6), round(oneyryield, 6), round(twoyryield, 6))
